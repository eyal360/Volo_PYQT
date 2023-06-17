import os, re, sys, shutil, webbrowser
import sqlite3, io, openpyxl, xlwt, pandas as pd, pandas.io.sql as sql
import time, threading, platform

import PIL
from PIL import Image

import emoji
from emoji import emojize
from datetime import date, datetime
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle, numbers
from openpyxl import *
from openpyxl import Workbook, writer
from numpy.lib import math
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication
from PyQt5.uic import loadUi
from PySide2 import QtCore, QtGui, QtWidgets
from PySide2.QtCore import (QCoreApplication, QPropertyAnimation, QDate, QDateTime, QMetaObject, QObject, QPoint, QRect,
                            QSize, QTime, QUrl, Qt, QEvent)
from PySide2.QtGui import (QBrush, QMovie, QColor, QConicalGradient, QCursor, QFont, QFontDatabase, QIcon,
                           QKeySequence, QLinearGradient, QPalette, QPainter, QPixmap, QRadialGradient)
from PySide2.QtWidgets import *

from ui_splash_screen import Ui_SplashScreen
from ui_customers_window import Ui_CustomersWindow
from ui_suppliers_window import Ui_SuppliersWindow
from ui_event_window import Ui_EventWindow
from ui_lead_window import Ui_LeadWindow
from ui_main_window import Ui_MainWindow
from ui_warning_window import Ui_WarningWindow

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ GLOBAL VARIABLES ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
DB = 'קבצי מערכת/' + 'database.db'
counter = 0

Music_Types = ['mp3', 'wav', 'band', 'asx', 'ram', 'zpl', 'm3u', 'pls', 'jam', 'gp*', 'mid', 'midi', 'ogg', 'vox',
               'voc', 'mp2', 'mp1',
               'ac3', 'amr', 'mpeg', 'gsm', 'wma', 'aac', 'mpc', 'ots', 'flac', 'raw', 'au', 'aiff']

Image_Types = ['jpg', 'jpeg', 'png', 'gif', 'ico', 'jng', 'jp2', 'jps', 'otb', 'pdn', 'tiff', 'vtf', 'xpm', 'mp4',
               'mov', 'avi', 'wmv',
               'wmp', 'ogg']

Document_Types = ['pdf', 'docx', 'pptx', 'ppt', 'xls', 'xlsx', 'doc', 'csv', 'docm', 'gdoc', 'mcw', 'log', 'txt',
                  'html', 'dot', 'AI',
                  'fm', 'gdraw', 'pdf', 'psd', 'pub', 'db', 'frm', 'ora', 'sqlite', 'wdb', 'lib', 'iso', 'cad', 'dmg',
                  'arc', 'bin',
                  'cab', 'xml', 'exe', 'gbp', 'gho', 'jar', 'izip', 'zip', '7zip', 'php', 'tar', 'tgz', 'wim']

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~ CUSTOMERS WINDOW VARIABLES ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

customers_tree_row = 0
current_customer = []
global customer_oid

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~ SUPPLIERS WINDOW VARIABLES ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

suppliers_tree_row = 0
current_supplier = []
global supplier_oid

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~ EVENTS AND LEADS WINDOWS ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
current_budget = []

events_dict = {}
events = []
leads_dict = {}
leads = []
event_pages = 1
lead_pages = 1
events_counter = 0
leads_counter = 0

payments_tree_row = 0
current_payment = []
global payment_oid

missions_tree_row = 0
current_mission = []
global mission_oid

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~ MAIN WINDOW VARIABLES ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
days_dict = {'Sunday': 'יום ראשון',
             'Monday': 'יום שני',
             'Tuesday': 'יום שלישי',
             'Wednesday': 'יום רביעי',
             'Thursday': 'יום חמישי',
             'Friday': 'יום שישי',
             'Saturday': 'יום שבת'}

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~ EXCEL VARIABLES ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
Mission_Cnt = 0

header = NamedStyle(name="header")
header.font = Font(name='Segoe UI Light', size=12, bold=True, color='000000')
header.border = Border(top=Side(border_style="thin"),
                       bottom=Side(border_style="thin"),
                       right=Side(border_style="thin"),
                       left=Side(border_style="thin")
                       )
header.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
header.alignment = Alignment(horizontal="center", vertical="center")
header.fill = PatternFill(start_color='CADBFF', fill_type='solid')

title = NamedStyle(name="title")
title.font = Font(name='Segoe UI Bold', size=14, bold=True, color='000000')
title.border = Border(right=Side(border_style="thin"),
                      left=Side(border_style="thin")
                      )
title.alignment = Alignment(horizontal="center", vertical="center")
title.fill = PatternFill(start_color='E7EBE6', fill_type='solid')

entry = NamedStyle(name="entry")
entry.font = Font(name='Calibri', size=12, bold=False, color='000000')
entry.border = Border(bottom=Side(border_style="thin"),
                      right=Side(border_style="thin"),
                      left=Side(border_style="thin")
                      )
entry.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
entry.alignment = Alignment(horizontal="center", vertical="center")


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

def backup_db():
    global DB
    one_time_FLAG = False
    run = True

    DB_Class = DataBase()
    backup_time = DB_Class.fetch_from_db('backup', 'all')
    if 'כניסה' in backup_time[0] or 'אף פעם' in backup_time[0]:
        one_time_FLAG = True
        wait_time = 0
    else:
        wait_time = backup_time[0].split(' ')[1]

    while run:
        if one_time_FLAG:
            run = False

        now = datetime.now()  # datetime object containing current date and time
        backup_db_name = 'Backup_DB ' + now.strftime("%d_%m_%Y %H_%M") + '.db'  # dd_mm_YY H_M

        folder_path = 'קבצי מערכת/'
        files = os.listdir(folder_path)

        for file_name in files:
            if 'Backup_DB' in file_name:
                old_Backup_DB = folder_path + "/" + file_name
                os.remove(old_Backup_DB)

        DB_path = DB
        exists_DB = DB.split('/')[1]
        new_DB_path = folder_path + "/" + backup_db_name
        if exists_DB in files:
            shutil.copyfile(DB_path, new_DB_path)

        print(f'backup - {backup_db_name} - has been made')
        time.sleep(int(wait_time) * 60)


def restart():
    SplashScreen.singleton = SplashScreen()


def make_date_standard(val):
    """
    convert date to the standard template,
    like  11-11-2021 -> 11.11.2021
    or 11-11-21 -> 11.11.2021
    """
    try:
        event_date = re.findall(r"[\w']+", val)
        if len(event_date[2]) == 2:
            event_date[2] = '20' + event_date[2]
        elif len(event_date[2]) == 3:
            event_date[2] = '2' + event_date[2]
        new_date = event_date[0] + '.' + event_date[1] + '.' + event_date[2]
    except:
        new_date = ''
    return new_date


def make_amount_readable(val):
    """
    send an amount and get it in readable form,
    like  10000 -> 10,000
    """
    if ',' in val:
        pass
    else:
        if len(str(val)) > 3:
            value = val[:-3] + ',' + val[-3:]
        elif len(str(val)) > 6:
            value = val[:-6] + ',' + val[-6:] + val[:-3] + ',' + val[-3:]
        else:
            value = val
    return value


def personal_folders(state, production_name):
    if state == 'create':
        for folder_path in ['הפקות/', 'לידים/']:
            sub_folder = folder_path + "/" + production_name
            if os.path.isdir(sub_folder):
                if not os.path.isdir(sub_folder + '/' + 'מסמכים'):
                    os.mkdir(sub_folder + '/' + 'מסמכים')
                if not os.path.isdir(sub_folder + '/' + 'תמונות וסרטונים'):
                    os.mkdir(sub_folder + '/' + 'תמונות וסרטונים')
                if not os.path.isdir(sub_folder + '/' + 'מוזיקה'):
                    os.mkdir(sub_folder + '/' + 'מוזיקה')

    elif state == 'delete':
        if production_name == 'all':
            for folder_path in ['הפקות/', 'לידים/']:
                for event in events:
                    sub_folder = folder_path + "/" + event[0]
                    if os.path.isdir(sub_folder):
                        shutil.rmtree(sub_folder, ignore_errors=True)
                for lead in leads:
                    sub_folder = folder_path + "/" + lead[0]
                    if os.path.isdir(sub_folder):
                        shutil.rmtree(sub_folder, ignore_errors=True)
        else:
            for folder_path in ['הפקות/', 'לידים/']:
                sub_folder = folder_path + "/" + production_name
                if os.path.isdir(sub_folder):
                    shutil.rmtree(sub_folder, ignore_errors=True)


def create_excel(*args):
    """
    Create excel report for each window
    args[0] - btn clicked for report name
    args[1] - event name(if needed)
    """
    global excel_image

    folder_path = 'הפקות/'
    if excel_image != '':
        img = openpyxl.drawing.image.Image(excel_image)

    if 'suppliers' in str(args[0]):
        wb_name = 'רשימת ספקים.xlsx'

        db_connect = sqlite3.connect(DB)
        df = sql.read_sql("SELECT * FROM suppliers", db_connect)
        df.to_excel(wb_name, 'רשימת ספקים')
        db_connect.close()

        # SET WORKSHEET
        wb = openpyxl.load_workbook(wb_name)
        ws = wb.active
        ws.sheet_view.rightToLeft = True

        # SET HEADER VALUES
        ws['A1'] = '#'
        ws['B1'] = 'שם חברה'
        ws['C1'] = 'כתובת החברה'
        ws['D1'] = 'איש קשר'
        ws['E1'] = 'טלפון'
        ws['F1'] = 'אימייל'
        ws['G1'] = 'פרטים נוספים'
        ws['H1'] = 'הסדר'

        # TITLE STYLE
        ws.insert_rows(1)
        ws.merge_cells('A1:H1')
        cell = ws.cell(row=1, column=1)
        ws.row_dimensions[1].height = 55
        cell.value = 'רשימת ספקים - Volo הפקות'
        cell.style = title
        if excel_image != '':
            img.anchor = 'H1'
            ws.add_image(img)

        # HEADERS STYLE
        header_row = ws[2]
        for cell in header_row:
            cell.style = header
        ws.row_dimensions[2].height = 20

        # ADD BORDERS TO ALL CELLS
        for row in range(3, ws.max_row + 1):
            for column in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=column)
                cell.style = entry

        wb.save(wb_name)
        wb.close()

        webbrowser.open(os.getcwd())

    if 'customers' in str(args[0]):
        wb_name = 'רשימת לקוחות.xlsx'
        db_connect = sqlite3.connect(DB)
        df = sql.read_sql("SELECT * FROM customers", db_connect)
        df.to_excel(wb_name, 'לקוחות')
        db_connect.close()

        # SET WORKSHEET
        wb = openpyxl.load_workbook(wb_name)
        ws = wb.active
        ws.sheet_view.rightToLeft = True

        # SET HEADER VALUES
        ws['A1'] = '#'
        ws['B1'] = 'שם האירוע'
        ws['C1'] = 'מועד'
        ws['D1'] = 'איש קשר'
        ws['E1'] = 'כתובת'
        ws['F1'] = 'טלפון'
        ws['G1'] = 'אימייל'
        ws['H1'] = 'חוזה'
        ws['I1'] = 'תעריף'
        ws['J1'] = 'אולם'
        ws['K1'] = 'כתובת האולם'
        ws['L1'] = 'סוג אירוע'
        ws['M1'] = 'ליד'

        # TITLE STYLE
        ws.insert_rows(1)
        ws.merge_cells('A1:M1')
        cell = ws.cell(row=1, column=1)
        ws.row_dimensions[1].height = 55
        cell.value = 'רשימת לקוחות - Volo הפקות'
        cell.style = title
        if excel_image != '':
            img.anchor = 'L1'
            ws.add_image(img)

        # HEADERS STYLE
        header_row = ws[2]
        for cell in header_row:
            cell.style = header
        ws.row_dimensions[2].height = 20

        # ADD BORDERS TO ALL CELLS
        for row in range(3, ws.max_row + 1):
            for column in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=column)
                cell.style = entry

        wb.save(wb_name)
        wb.close()

        webbrowser.open(os.getcwd())

    if 'payments' in str(args[0]) or 'event' in str(args[0]):
        sheet_name = 'ניהול תקציב'
        wb_name = folder_path + "/" + args[1] + '/' + sheet_name + ' - ' + args[1] + '.xlsx'

        db_connect = sqlite3.connect(DB)
        df = sql.read_sql(f"SELECT * FROM payments WHERE event_name ='{args[1]}' ORDER BY category DESC", db_connect)
        df.to_excel(wb_name, sheet_name)
        db_connect.close()

        # SET WORKSHEET
        wb = openpyxl.load_workbook(wb_name)
        ws = wb[sheet_name]
        ws.sheet_view.rightToLeft = True
        ws.delete_cols(2)

        # SET HEADER VALUES
        ws['A1'] = '#'
        ws['B1'] = 'קטגוריה (% מסך ההוצאות)'
        ws['C1'] = 'שירות'
        ws['D1'] = 'ספק'
        ws['E1'] = 'עלות'
        ws['F1'] = 'מקדמה'
        ws['G1'] = 'מקדמה שולמה'
        ws['H1'] = 'טיפ'
        ws['I1'] = 'פרטים נוספים'

        # TITLE STYLE
        ws.insert_rows(1)
        ws.merge_cells('A1:I1')
        cell = ws.cell(row=1, column=1)
        ws.row_dimensions[1].height = 55
        cell.value = sheet_name + '-' + args[1] + '-' + 'Volo הפקות'
        cell.style = title
        if excel_image != '':
            img.anchor = 'J1'
            ws.add_image(img)

        # SUB-TITLE STYLE
        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("SELECT * FROM customers WHERE event_name=:event_name",
                       {'event_name': args[1]
                        })
        record = cursor.fetchone()
        event_type = record[10]
        location = record[8]
        cursor.close()
        db_connect.close()

        ws.insert_rows(2)
        ws.merge_cells('A2:I2')
        cell = ws.cell(row=2, column=1)
        ws.row_dimensions[2].height = 30
        cell.value = event_type + ' ' + location
        cell.style = title

        # HEADERS STYLE
        header_row = ws[3]
        for cell in header_row:
            cell.style = header
        ws.row_dimensions[3].height = 20

        # FORMAT NUMBER VALUES TO FLOAT AND INTEGER FOR SUM

        # VARIABLES DEFINITION
        skip_cnt = 3  # how many lines to skip before starting to convert
        current_row = skip_cnt
        total_payments = 0
        total_gifts = 0
        total_tips = 0
        for row in range(1, ws.max_row - 1):
            current_row += 1
            if ws[f'E{current_row}'].value:
                ws[f'E{current_row}'].value = float(ws[f'E{current_row}'].value)
                ws[f'E{current_row}'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                if ws[f'B{current_row}'].value == 'מתנות':
                    ws[f'E{current_row}'].value = (0 - float(ws[f'E{current_row}'].value))
                    total_gifts += abs(ws[f'E{current_row}'].value)
                total_payments += ws[f'E{current_row}'].value

            if ws[f'F{current_row}'].value:
                ws[f'F{current_row}'].value = float(ws[f'F{current_row}'].value)
                ws[f'F{current_row}'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            if ws[f'H{current_row}'].value:
                ws[f'H{current_row}'].value = int(ws[f'H{current_row}'].value)
                ws[f'H{current_row}'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                total_tips += ws[f'H{current_row}'].value

        # MERGING SAME CATEGORIES

        # VARIABLES DEFINITION
        start_row = 5
        i = 1

        for row in range(start_row, ws.max_row + 1):  # from 4 to 21
            if ws[f'B{row}'].value == ws[f'B{row - 1}'].value:
                i += 1

            elif i != 1:
                ws.merge_cells(f'B{row - i}:B{row - 1}')
                i = 1

        # ADD SUMMARY VALUES TO PAYMENT, TIPS AND BUDGET

        # PAYMENTS
        ws.insert_rows(ws.max_row)
        ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')
        cell = ws.cell(row=ws.max_row, column=1)
        ws.row_dimensions[2].height = 25
        cell.value = 'סה"כ הוצאות'
        cell.style = title

        total_payments_cell = ws.cell(row=ws.max_row, column=5)
        total_payments_cell.value = float(total_payments)
        total_payments_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        total_payments_cell.style = header

        # TIPS
        ws.merge_cells(f'F{ws.max_row}:G{ws.max_row}')
        cell = ws.cell(row=ws.max_row, column=6)
        ws.row_dimensions[2].height = 25
        cell.value = 'סה"כ טיפ'
        cell.style = title

        total_tip_cell = ws.cell(row=ws.max_row, column=8)
        total_tip_cell.value = float(total_tips)
        total_tip_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        total_tip_cell.style = header

        # BUDGET
        ws.merge_cells(f'A{ws.max_row + 2}:D{ws.max_row + 2}')
        cell = ws.cell(row=ws.max_row, column=1)
        ws.row_dimensions[2].height = 25
        cell.value = 'תקציב האירוע'
        cell.style = title

        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("SELECT budget FROM budgets WHERE event_name=:event_name",
                       {'event_name': args[1]
                        })
        record = cursor.fetchone()

        if record and record[0] != '':  # if record not empty
            budget = record[0]
        else:
            budget = 'לא סוכם'

        total_payments_cell = ws.cell(row=ws.max_row, column=5)
        total_payments_cell.value = float(total_payments)
        total_payments_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        total_payments_cell.style = header

        total_budget_cell = ws.cell(row=ws.max_row, column=5)
        try:
            total_budget_cell.value = float(budget)
        except:
            total_budget_cell.value = budget
        total_budget_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        total_budget_cell.style = header

        # BUDGET STATUS
        ws.merge_cells(f'A{ws.max_row + 2}:D{ws.max_row + 2}')
        cell = ws.cell(row=ws.max_row, column=1)
        ws.row_dimensions[2].height = 25
        cell.style = title

        budget_status_cell = ws.cell(row=ws.max_row, column=5)
        budget_status_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        budget_status_cell.style = header

        try:
            if int(budget) - int(total_payments) >= 0:  # IF UNDER BUDGET
                cell.value = 'יתרה תקציבית'
                budget_status_cell.value = int(budget) - int(total_payments)
                budget_status_cell.fill = PatternFill(start_color='589A0A', fill_type='solid')

            else:  # IF OVER BUDGET
                cell.value = 'חורג מהתקציב בסך'
                cell.fill = PatternFill(start_color='E55261', fill_type='solid')
                budget_status_cell.value = abs(int(budget) - int(total_payments))
                budget_status_cell.fill = PatternFill(start_color='DF283A', fill_type='solid')
                budget_status_cell.font = Font(name='Segoe UI Light', size=12, bold=True, color='FFFFFF')

        except:
            cell.value = 'יתרה תקציבית'
            budget_status_cell.value = 0

        # ADD BORDERS TO ALL CELLS
        for row in range(4, ws.max_row - 5):
            for column in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=column)
                cell.style = entry

        # ADD DETAILS ABOUT PAYMENTS VALUE-PERCENTAGE FROM THE TOTAL COST

        # VARIABLES DEFINITION
        start_row = 4
        starting_cell_flag = False
        gift_flag = False
        percentage = 0
        total_payments += total_gifts

        for row in range(start_row, ws.max_row - 5):  # from 4 to 21
            payment_cell = ws[f"E{row}"]
            category_cell = ws[f"B{row}"]
            next_category_cell = ws[f"B{row + 1}"]

            if category_cell.value == 'מתנות' or gift_flag == True:
                payment_cell.fill = PatternFill(start_color='78BD7A', fill_type='solid')
                gift_flag = True
                if type(next_category_cell).__name__ == 'Cell':
                    gift_flag = False

            elif type(category_cell).__name__ == 'Cell' and type(next_category_cell).__name__ == 'Cell':
                percentage = 0
                if payment_cell.value:
                    percentage = (float(payment_cell.value) / total_payments) * 100
                category_cell.value += ' (' + str(round(percentage, 2)) + '%' + ') '

            elif type(category_cell).__name__ == 'Cell' and type(next_category_cell).__name__ == 'MergedCell':
                starting_cell_flag = True
                starting_cell = category_cell
                if payment_cell.value:
                    percentage = (float(payment_cell.value) / total_payments) * 100

            elif type(category_cell).__name__ == 'MergedCell' and type(next_category_cell).__name__ == 'MergedCell':
                if payment_cell.value:
                    percentage += (float(payment_cell.value) / total_payments) * 100

            elif type(category_cell).__name__ == 'MergedCell' and type(next_category_cell).__name__ == 'Cell':
                if starting_cell_flag == True:
                    starting_cell_flag = False
                if payment_cell.value:
                    percentage += (float(payment_cell.value) / total_payments) * 100
                starting_cell.value += ' (' + str(round(percentage, 2)) + '%' + ') '

        wb.save(wb_name)
        wb.close()

        webbrowser.open(os.path.realpath(folder_path + args[1]))

    if 'missions' in str(args[0]) or 'event' in str(args[0]):
        sheet_name = 'משימות'
        wb_name = folder_path + "/" + args[1] + '/' + sheet_name + ' - ' + args[1] + '.xlsx'

        db_connect = sqlite3.connect(DB)
        df = sql.read_sql(f"SELECT * FROM missions WHERE event_name='{args[1]}' ORDER BY mission_due", db_connect)
        df.to_excel(wb_name, sheet_name)
        db_connect.close()

        # SET WORKSHEET
        wb = openpyxl.load_workbook(wb_name)
        ws = wb[sheet_name]
        ws.sheet_view.rightToLeft = True
        ws.delete_cols(2)

        # SET HEADER VALUES
        ws['A1'] = '#'
        ws['B1'] = 'משימה'
        ws['C1'] = 'לביצוע עד'
        ws['D1'] = 'סטאטוס'
        ws['E1'] = 'הערות'

        # TITLE STYLE
        ws.insert_rows(1)
        ws.merge_cells('A1:E1')
        cell = ws.cell(row=1, column=1)
        ws.row_dimensions[1].height = 25
        cell.value = sheet_name + ' - ' + args[1]
        cell.style = title
        if excel_image != '':
            img.anchor = 'F1'
            ws.add_image(img)

        # HEADERS STYLE
        header_row = ws[2]
        for cell in header_row:
            cell.style = header
        ws.row_dimensions[2].height = 20

        # ADD BORDERS TO ALL CELLS AND COLORS TO STATUS CELLS

        status_color = {'בתהליך': 'FFDF80',
                        'בוצע': 'A5D46A',
                        'לא רלוונטי': 'E38C8C'}

        for row in range(3, ws.max_row + 1):
            for column in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=column)
                cell.style = entry
                try:
                    cell.fill = PatternFill(start_color=status_color[cell.value], fill_type='solid')
                except Exception as e:
                    pass

        wb.save(wb_name)
        wb.close()

        webbrowser.open(os.path.realpath(folder_path + args[1]))


# DATA BASE
class DataBase():
    global DB

    def __init__(self):
        db_path = 'קבצי מערכת/'
        if not os.path.isdir(db_path):
            os.mkdir(db_path)

    def run_DB_backup(self):
        backup_time = self.fetch_from_db('backup', 'all')
        if backup_time == None or backup_time == '':
            pass
        elif backup_time[1] == 1:
            threading.Thread(target=backup_db).start()
        return

    def create_tables(self):
        """
            This function created all DB tables if not exist
        """
        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("""CREATE TABLE IF NOT EXISTS customers (
                                event_name text,
                                event_date text,
                                contact_name text,
                                contact_address text,
                                contact_phone text,
                                contact_email text,
                                contract text,
                                rate text,
                                location text,
                                location_address text,
                                place text,
                                lead text
                                )""")
        # Create Suppliers table
        cursor.execute("""CREATE TABLE IF NOT EXISTS suppliers (
                                company text,
                                company_address text,
                                contact_name text,
                                contact_phone text,
                                contact_email text,
                                details text,
                                contract text
                                )""")
        # Create Payments table
        cursor.execute("""CREATE TABLE IF NOT EXISTS payments (
                                event_name text,
                                category text,
                                service text,
                                supplier text,
                                price text,
                                advance text,
                                advance_status text,
                                tip text,
                                details text
                                )""")
        # Create Missions table
        cursor.execute("""CREATE TABLE IF NOT EXISTS missions (
                                event_name text,
                                mission text,
                                mission_due text,
                                status text,
                                notes text
                                )""")
        # Create Budgets table
        cursor.execute("""CREATE TABLE IF NOT EXISTS budgets (
                                event_name text,
                                budget text
                                )""")
        # Create Categories table
        cursor.execute("""CREATE TABLE IF NOT EXISTS categories (
                                        category text
                                        )""")
        # Create Places table
        cursor.execute("""CREATE TABLE IF NOT EXISTS places (
                                        place text
                                        )""")
        # Create Places backup
        cursor.execute("""CREATE TABLE IF NOT EXISTS backup (
                                                time text
                                                )""")
        cursor.close()
        db_connect.close()

    def fetch_events_and_leads(self):
        global events
        global leads
        global event_pages
        global lead_pages
        global events_counter
        global leads_counter

        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("SELECT DISTINCT event_name, event_date FROM customers WHERE lead=:lead",
                       {'lead': 'לא'
                        })
        events = cursor.fetchall()
        cursor.execute("SELECT DISTINCT event_name, event_date FROM customers WHERE lead=:lead ",
                       {'lead': 'כן'
                        })
        leads = cursor.fetchall()

        event_pages = math.ceil(len(events) / 9)
        lead_pages = math.ceil(len(leads) / 9)

        cursor.close()
        db_connect.close()

    def insert_into_db(self, *args, **kwargs):
        """
            This function insert info to db table
            syntax - (db_name, info_dict)
        """
        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()

        if args[0] == 'customers':
            cursor.execute("INSERT INTO customers VALUES (:event_name, :event_date,"
                           " :contact_name, :contact_address, :contact_phone,"
                           " :contact_email, :contract, :rate, :location, :location_address, :place, :lead)",
                           {'event_name': args[1]['event_name'],
                            'event_date': args[1]['event_date'],
                            'contact_name': args[1]['contact_name'],
                            'contact_address': args[1]['contact_address'],
                            'contact_phone': args[1]['contact_phone'],
                            'contact_email': args[1]['contact_email'],
                            'contract': args[1]['contract'],
                            'rate': args[1]['rate'],
                            'location': args[1]['location'],
                            'location_address': args[1]['location_address'],
                            'place': args[1]['place'],
                            'lead': args[1]['lead']
                            })

        elif args[0] == 'suppliers':
            cursor.execute("INSERT INTO suppliers VALUES ("
                           ":company, :company_address, :contact_name, :contact_phone, :contact_email, :details, :contract)",
                           {'company': args[1]['company'],
                            'company_address': args[1]['company_address'],
                            'contact_name': args[1]['contact_name'],
                            'contact_phone': args[1]['contact_phone'],
                            'contact_email': args[1]['contact_email'],
                            'details': args[1]['details'],
                            'contract': args[1]['contract']
                            })

        elif args[0] == 'payments':
            cursor.execute("INSERT INTO payments VALUES ("
                           ":event_name, :category, :service, :supplier, :price, :advance,"
                           " :advance_status, :tip, :details)",
                           {'event_name': args[1]['event_name'],
                            'category': args[1]['category'],
                            'service': args[1]['service'],
                            'supplier': args[1]['supplier'],
                            'price': args[1]['price'],
                            'advance': args[1]['advance'],
                            'advance_status': args[1]['advance_status'],
                            'tip': args[1]['tip'],
                            'details': args[1]['details']
                            })

        elif args[0] == 'missions':
            cursor.execute("INSERT INTO missions VALUES ("
                           ":event_name, :mission, :mission_due, :status, :notes)",
                           {'event_name': args[1]['event_name'],
                            'mission': args[1]['mission'],
                            'mission_due': args[1]['mission_due'],
                            'status': args[1]['status'],
                            'notes': args[1]['notes']
                            })
        elif args[0] == 'budgets':
            cursor.execute("INSERT INTO budgets VALUES ("
                           ":event_name, :budget)",
                           {'event_name': args[1]['event_name'],
                            'budget': args[1]['budget']
                            })
        elif args[0] == 'categories':
            cursor.execute("INSERT INTO categories VALUES ("
                           ":category)",
                           {'category': args[1]
                            })
        elif args[0] == 'places':
            cursor.execute("INSERT INTO places VALUES ("
                           ":place)",
                           {'place': args[1]
                            })

        elif args[0] == 'backup':
            cursor.execute("INSERT INTO backup VALUES ("
                           ":time)",
                           {'time': args[1]
                            })

        db_connect.commit()
        cursor.close()
        db_connect.close()

    def fetch_from_db(self, *args, **kwargs):
        """
            This function fetch info from db table
            syntax for customers - (db_name, all/one, event_name, contact_name)
            syntax for suppliers - (db_name, all/one, company, contact_name)
            syntax for payments - (db_name, all/one, event_name(ALL), supplier(ONE), service (ONE))
            syntax for missions - (db_name, all/one, event_name(ALL), mission(ONE))
            syntax for budgets - (db_name, all/one, event_name(ONE))
            syntax for categories - (db_name)
            syntax for places - (db_name)
            syntax for time - (db_name, 'all')
        """
        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()

        if args[0] == 'customers':
            if args[1] == 'all':
                cursor.execute("SELECT *,oid FROM customers")
                records = cursor.fetchall()
                cursor.close()
                db_connect.close()
                return records


            elif args[1] == 'one':
                cursor.execute(
                    "SELECT *,oid FROM customers WHERE event_name=:event_name AND contact_name=:contact_name",
                    {'event_name': args[2],
                     'contact_name': args[3]
                     })
                records = cursor.fetchall()
                cursor.close()
                db_connect.close()
                return records

        elif args[0] == 'suppliers':
            if args[1] == 'all':
                cursor.execute("SELECT *,oid FROM suppliers")
                records = cursor.fetchall()
                cursor.close()
                db_connect.close()
                return records

            elif args[1] == 'one':
                cursor.execute("SELECT *,oid FROM suppliers WHERE company=:company AND contact_name=:contact_name",
                               {'company': args[2],
                                'contact_name': args[3]
                                })
                records = cursor.fetchall()
                cursor.close()
                db_connect.close()
                return records

        elif args[0] == 'payments':
            if args[1] == 'all':
                cursor.execute("SELECT *,oid FROM payments WHERE event_name=:event_name",
                               {'event_name': args[2]
                                })
                records = cursor.fetchall()
                cursor.close()
                db_connect.close()
                return records

            if args[1] == 'one':
                cursor.execute("SELECT *,oid FROM payments WHERE "
                               "event_name=:event_name AND supplier=:supplier AND service=:service",
                               {'event_name': args[2],
                                'supplier': args[3],
                                'service': args[4]
                                })
                records = cursor.fetchall()
                cursor.close()
                db_connect.close()
                return records

        elif args[0] == 'missions':
            if args[1] == 'all':
                cursor.execute("SELECT * FROM missions WHERE event_name=:event_name",
                               {'event_name': args[2]
                                })
                records = cursor.fetchall()
                cursor.close()
                db_connect.close()
                return records

            if args[1] == 'one':
                cursor.execute("SELECT *,oid FROM missions WHERE event_name=:event_name AND mission=:mission",
                               {'event_name': args[2],
                                'mission': args[3]
                                })
                records = cursor.fetchall()
                cursor.close()
                db_connect.close()
                return records

        if args[0] == 'budgets':
            if args[1] == 'all':
                cursor.execute("SELECT * FROM budgets")
                records = cursor.fetchall()
                cursor.close()
                db_connect.close()
                return records

            if args[1] == 'one':
                cursor.execute("SELECT * FROM budgets WHERE event_name=:event_name",
                               {'event_name': args[2]
                                })
                records = cursor.fetchone()
                cursor.close()
                db_connect.close()
                return records

        if args[0] == 'categories':
            cursor.execute("SELECT DISTINCT category FROM categories")
            records = cursor.fetchall()
            cursor.close()
            db_connect.close()
            return records

        if args[0] == 'places':
            cursor.execute("SELECT DISTINCT place FROM places")
            records = cursor.fetchall()
            cursor.close()
            db_connect.close()
            return records

        if args[0] == 'backup':
            if args[1] == 'all':
                cursor.execute("SELECT *,oid FROM backup")
                records = cursor.fetchone()
                cursor.close()
                db_connect.close()
                return records

    def update_db(self, *args, **kwargs):
        """
            This function fetch info from db table
            syntax for customers - (db_name)
            syntax for suppliers - (db_name)
            syntax for payments - (db_name)
            syntax for missions - (db_name)
            syntax for budgets - (db_name)
            syntax for backup - (db_name, value, time_oid)
        """
        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()

        if args[0] == 'customers':
            global current_customer

            event_name, event_date, contact_name, contact_address, contact_phone, contact_email, contract, \
            rate, location, location_address, place, lead, oid = current_customer  ## set up variables from tuple
            cursor.execute("""UPDATE customers SET 
                            event_name = :event_name,
                            event_date = :event_date,
                            contact_name = :contact_name,
                            contact_address = :contact_address,
                            contact_phone = :contact_phone,
                            contact_email = :contact_email,
                            contract = :contract,
                            rate = :rate ,
                            location = :location,
                            location_address = :location_address,
                            place = :place,
                            lead = :lead
                            WHERE oid = :oid""",
                           {'event_name': event_name,
                            'event_date': event_date,
                            'contact_name': contact_name,
                            'contact_address': contact_address,
                            'contact_phone': contact_phone,
                            'contact_email': contact_email,
                            'contract': contract,
                            'rate': rate,
                            'location': location,
                            'location_address': location_address,
                            'place': place,
                            'lead': lead,
                            'oid': oid
                            })
            current_customer = []

        elif args[0] == 'suppliers':
            global current_supplier

            company, company_address, contact_name, contact_phone, contact_email, details, \
            contract, customer_oid = current_supplier
            cursor.execute("""UPDATE suppliers SET 
                            company = :company,
                            company_address = :company_address,
                            contact_name = :contact_name,
                            contact_phone = :contact_phone,
                            contact_email = :contact_email,
                            details = :details,
                            contract = :contract
                            WHERE oid = :oid""",
                           {'company': company,
                            'company_address': company_address,
                            'contact_name': contact_name,
                            'contact_phone': contact_phone,
                            'contact_email': contact_email,
                            'details': details,
                            'contract': contract,
                            'oid': customer_oid
                            })

            current_supplier = []

        elif args[0] == 'payments':
            global current_payment

            event_name, category, service, supplier, price, advance, \
            advance_status, tip, details, payment_oid = current_payment
            cursor.execute("""UPDATE payments SET 
                            event_name = :event_name,
                            category = :category,
                            service = :service,
                            supplier = :supplier,
                            price = :price,
                            advance = :advance,
                            advance_status = :advance_status,
                            tip = :tip,
                            details = :details
                            WHERE oid = :oid""",
                           {'event_name': event_name,
                            'category': category,
                            'service': service,
                            'supplier': supplier,
                            'price': price,
                            'advance': advance,
                            'advance_status': advance_status,
                            'tip': tip,
                            'details': details,
                            'oid': payment_oid
                            })

            current_payment = []

        elif args[0] == 'missions':
            global current_mission
            event_name, mission, mission_due, status, notes, mission_oid = current_mission
            cursor.execute("""UPDATE missions SET
                            event_name = :event_name,
                            mission = :mission,
                            mission_due = :mission_due,
                            status = :status,
                            notes = :notes
                            WHERE oid = :oid""",
                           {'event_name': event_name,
                            'mission': mission,
                            'mission_due': mission_due,
                            'status': status,
                            'notes': notes,
                            'oid': mission_oid
                            })
            current_mission = []

        elif args[0] == 'budgets':
            global current_budget
            event_name, budget = current_budget
            cursor.execute("""UPDATE budgets SET
                            event_name = :event_name,
                            budget = :budget
                            WHERE event_name = :event_name""",
                           {'event_name': event_name,
                            'budget': budget
                            })
            current_budget = []

        elif args[0] == 'backup':
            cursor.execute("""UPDATE backup SET 
                            time = :time 
                            WHERE oid = :oid""",
                           {'time': args[1],
                            'oid': args[2]
                            })

        db_connect.commit()
        cursor.close()
        db_connect.close()

    def delete_from_db(self, *args, **kwargs):
        """
        This function fetch info from db table
        syntax for customers - (db_name, , all/one, event_name(ONE), full_name(ONE))
        syntax for suppliers - (db_name, all/one, company(ONE))
        syntax for payments - (db_name, all/one, event_name(ALL), supplier(ONE), service (ONE)
        syntax for missions - (db_name, all/one, event_name(ALL), mission(ONE))
        syntax for backup - (db_name, all)
        """
        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()

        if args[0] == 'customers':
            if args[1] == 'all':
                cursor.execute("DELETE FROM customers")

            elif args[1] == 'one':
                cursor.execute("DELETE FROM customers WHERE event_name=:event_name AND contact_name=:contact_name",
                               {'event_name': args[2],
                                'contact_name': args[3]
                                })

        elif args[0] == 'suppliers':
            if args[1] == 'all':
                cursor.execute("DELETE FROM suppliers")

            elif args[1] == 'one':
                cursor.execute("DELETE FROM suppliers WHERE company=:company AND contact_name=:contact_name",
                               {'company': args[2],
                                'contact_name': args[3]
                                })

        elif args[0] == 'payments':
            if args[1] == 'all':
                cursor.execute("DELETE FROM payments WHERE event_name=:event_name",
                               {'event_name': args[2]
                                })

            elif args[1] == 'one':
                cursor.execute("DELETE FROM payments WHERE "
                               "event_name=:event_name AND supplier=:supplier AND service=:service",
                               {'event_name': args[2],
                                'service': args[3],
                                'supplier': args[4]
                                })

        elif args[0] == 'missions':
            print('in missions')
            if args[1] == 'all':
                cursor.execute("DELETE FROM missions WHERE event_name=:event_name",
                               {'event_name': args[2],
                                })

            elif args[1] == 'one':
                cursor.execute("DELETE FROM missions WHERE event_name=:event_name AND mission=:mission",
                               {'event_name': args[2],
                                'mission': args[3]
                                })

        elif args[0] == 'backup':
                if args[1] == 'all':
                    cursor.execute("DELETE FROM backup")

        db_connect.commit()
        cursor.close()
        db_connect.close()


# DRAG AND DROP
class ListBoxWidget(QListWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls:
            event.accept()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(Qt.CopyAction)
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(Qt.CopyAction)
            event.accept()

            links = []
            for url in event.mimeData().urls():
                if url.isLocalFile():
                    links.append(str(url.toLocalFile()))
                else:
                    links.append(str(url.toString()))
        else:
            event.ignore()

        for item in links:
            self.add_file(item)

    def add_file(self, file):
        folder_path = 'הפקות/'
        file_name = file.split("/")[-1]
        file_type = file.split("/")[-1].split(".")[-1]
        if file_type in Image_Types:
            shutil.copyfile(file, folder_path + "/" + self.event_name + '/' + 'תמונות וסרטונים' + '/' + file_name)
            self.addItem(file_name.split(".")[0] + ' - נוסף בהצלחה')

        elif file_type in Document_Types:
            shutil.copyfile(file, folder_path + "/" + self.event_name + '/' + 'מסמכים' + '/' + file_name)
            self.addItem(file_name.split(".")[0] + ' - נוסף בהצלחה')

        elif file_type in Music_Types:
            shutil.copyfile(file, folder_path + "/" + self.event_name + '/' + 'מוזיקה' + '/' + file_name)
            self.addItem(file_name.split(".")[0] + ' - נוסף בהצלחה')

        else:
            try:
                shutil.copyfile(file, folder_path + "/" + self.event_name + '/' + file_name)
                self.addItem(file_name.split(".")[0] + ' - נוסף בהצלחה')
            except:
                pass


# MAIN WINDOW
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.event_window = EventWindow()
        self.lead_window = LeadWindow()
        self.setWindowIcon(QtGui.QIcon('Pheonix.ico'))
        self.fonts()

        # MAIN LABELS
        self.current_date()
        self.backup_info()

        # MAIN BUTTONS
        self.customers_window = CustomersWindow()
        self.ui.clients_button.clicked.connect(self.customers_window.pop)
        self.suppliers_window = SuppliersWindow()
        self.ui.suppliers_button.clicked.connect(self.suppliers_window.pop)
        self.ui.upper_frame_logo_btn.clicked.connect(self.restart)
        self.ui.setting_btn.clicked.connect(self.settings_menu)

        # EVENTS AND LEADS - STACKED WIDGETS
        self.ui.stacked_events_and_leads.setCurrentWidget(self.ui.home_page)
        self.ui.events_button.clicked.connect(self.stacked_events)
        self.ui.leads_button.clicked.connect(self.stacked_leads)
        self.ui.upper_frame_header_btn.clicked.connect(self.stacked_homepage)
        self.ui.upper_frame_logo_btn.clicked.connect(self.stacked_homepage)

        # HOME BUTTONS GIF
        self.ui.logo_browse_btn.clicked.connect(self.excel_logo_state)
        self.ui.reset_logo_btn.clicked.connect(self.excel_logo_state)
        self.excel_logo_state()

        ## BACKUP
        self.ui.backup_Setup_btn.clicked.connect(self.show_backup_settings)
        self.ui.backup_options_combobox.currentTextChanged.connect(self.setup_backup_details)
        self.ui.restore_backup_btn.clicked.connect(self.restore_backup)
        self.ui.restart_for_backup_btn.clicked.connect(self.restart)

        # self.ui.playGIF_btn.clicked.connect(self.play_GIF)
        # self.ui.pauseGIF_btn.clicked.connect(self.pause_GIF)
        # self.ui.nextGIF_btn.clicked.connect(self.next_GIF)
        #
        # self.ui.GIF_label.setMinimumSize(QtCore.QSize(400, 400))
        # self.ui.GIF_label.setMaximumSize(QtCore.QSize(800, 800))
        # self.movie = QMovie(gif_dict[gif_id])
        # self.ui.GIF_label.setMovie(self.movie)
        # self.movie.start()

        # EVENTS - WIDGETS
        self.ui.previous_events_btn.hide()
        if event_pages == 1:
            self.ui.next_events_btn.hide()
        self.ui.next_events_btn.clicked.connect(self.next_events_page_btn)
        self.ui.previous_events_btn.clicked.connect(self.previous_events_page_btn)
        self.event_pages_and_buttons()

        # LEADS - WIDGETS
        self.ui.previous_leads_btn.hide()
        if lead_pages == 1:
            self.ui.next_leads_btn.hide()
        self.ui.next_leads_btn.clicked.connect(self.next_leads_page_btn)
        self.ui.previous_leads_btn.clicked.connect(self.previous_leads_page_btn)
        self.lead_pages_and_buttons()

        # SOCIAL NETWORKS BUTTONS
        self.ui.FB_btn.clicked.connect(lambda: self.btn_to_web('fb'))
        self.ui.whatsapp_btn.clicked.connect(lambda: self.btn_to_web('whatsapp'))
        self.ui.instagram_btn.clicked.connect(lambda: self.btn_to_web('instagram'))
        self.ui.gmail_btn.clicked.connect(lambda: self.btn_to_web('gmail'))
        self.ui.calender_btn.clicked.connect(lambda: self.btn_to_web('calender'))

    ## SETTINGS
    def backup_info(self):
        folder_path = 'קבצי מערכת/'
        files = os.listdir(folder_path)
        raw_date = ''
        raw_time = ''

        for file in files:
            if 'Backup_DB' in file:
                raw_date = file.split(' ')[1]
                raw_time = file.split(' ')[2].split('.')[0]

        if raw_date != '' and raw_time != '':
            day_time = raw_time.split('_')[0] + ':' + raw_time.split('_')[1]
            day_date = raw_date.split('_')[0] + '/' + raw_date.split('_')[1] + '/' + raw_date.split('_')[2]
            label = f'גיבוי אחרון נוצר בתאריך {day_date} בשעה {day_time}'
        else:
            label = ''

        self.ui.backup_info_label.setText(label)

    def current_date(self):
        day_name = date.today().strftime("%A")
        day_date = date.today().strftime("%d/%m/%Y")
        self.ui.leads_date_label.setText(day_date + ' | ' + days_dict[day_name])

    def settings_menu(self):
        self.ui.stacked_events_and_leads.setCurrentWidget(self.ui.home_page)
        if self.ui.logo_browse_btn.isHidden():
            self.ui.quote_label.hide()
            self.ui.logo_browse_btn.show()
            self.ui.reset_logo_btn.show()
            self.ui.backup_Setup_btn.show()
            self.ui.restore_backup_btn.show()
            self.ui.line.show()
        else:
            self.ui.quote_label.show()
            self.ui.logo_browse_btn.hide()
            self.ui.reset_logo_btn.hide()
            self.ui.backup_Setup_btn.hide()
            self.ui.restore_backup_btn.hide()
            self.ui.line.hide()

    def restart(self):
        self.close()
        restart()
        # SplashScreen.singleton = SplashScreen()

    def fonts(self):
        global font100
        font100 = QFont()
        font100.setFamily(u"Segoe UI Semilight")
        font100.setPointSize(15)
        font100.setBold(False)
        font100.setItalic(False)
        font100.setWeight(7)

        global font101
        font101 = QFont()
        font101.setFamily(u"Segoe UI Semilight")
        font101.setPointSize(15)
        font101.setBold(False)
        font101.setItalic(False)
        font101.setWeight(7)

    def show_backup_settings(self):
        self.backup_time = DataBase.fetch_from_db(self, 'backup', 'all')
        self.btn_cnt = 0
        if self.backup_time == None or self.backup_time == '':
            self.ui.backup_options_combobox.setCurrentText('אף פעם')
        elif not self.backup_time[1] == 1:
            self.ui.backup_options_combobox.setCurrentText('אף פעם')
        else:
            self.ui.backup_options_combobox.setCurrentText(self.backup_time[0])

        if self.ui.backup_options_combobox.isHidden():
            self.ui.backup_freq_label.show()
            self.ui.backup_options_combobox.show()
        else:
            self.ui.backup_freq_label.hide()
            self.ui.backup_options_combobox.hide()

    def setup_backup_details(self, value):
        self.btn_cnt += 1
        self.backup_time = DataBase.fetch_from_db(self, 'backup', 'all')

        if self.backup_time != None:
            if self.backup_time[1] != 1:
                DataBase.insert_into_db(self, 'backup', value)
            else:
                DataBase.update_db(self, 'backup', value, self.backup_time[1])
                if self.btn_cnt > 1:
                    self.ui.restart_backup_warning.show()
                    self.ui.restart_for_backup_btn.show()
        else:
            DataBase.insert_into_db(self, 'backup', value)

    def restore_backup(self):
        folder_path = 'קבצי מערכת/'
        files = os.listdir(folder_path)
        old_DB = ''
        new_DB = ''

        for file in files:
            if 'database' in file:
                old_DB = folder_path + "/" + file
            elif 'DB' in file:
                new_DB = folder_path + "/" + file

        if old_DB != '' and new_DB != '':
            os.remove(old_DB)
            shutil.copyfile(new_DB, old_DB)
            self.restart()

    def excel_logo_state(self):
        global excel_image
        try:
            btn_name = str(self.sender())
        except:
            btn_name = ''
        folder_path = 'קבצי מערכת/'
        files = os.listdir(folder_path)
        delete_file = ''
        excel_image = ''

        for file in files:  # check if there is already an image file
            file_type = file.split(".")[-1]
            if file_type in Image_Types:
                delete_file = file

        if delete_file != '':
            excel_image = folder_path + '/' + delete_file
            self.ui.logo_browse_btn.setText('קיים לוגו, לחץ לשינוי')

        if 'browse' in btn_name:
            file = QFileDialog.getOpenFileNames(self, 'בחר קובץ', os.getcwd())

            if file[1] != '':  # check if the new selected file is an image file
                file_name = file[0][0].split("/")[-1]
                file_type = file[0][0].split("/")[-1].split(".")[-1]
                if len(file[0]) == 1:
                    if file_type in Image_Types:
                        if delete_file != '':
                            os.remove(folder_path + '/' + delete_file)  # delete the old image file
                        image = Image.open(file[0][0])
                        image.thumbnail(size=(90, 74))
                        image.save(folder_path + '/' + file_name, optimize=True, quality=65)
                        excel_image = folder_path + '/' + file_name

                        self.ui.logo_browse_btn.setText('נבחר לוגו, לחץ לשינוי')
                        self.ui.logo_browse_btn.setStyleSheet(u"QPushButton {\n"
                                                              "	background-color: rgb(101,180,115);	\n"
                                                              "	color: rgb(255, 255, 255);\n"
                                                              "	border: 1px solid #090206;\n"
                                                              "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                              "	border-radius: 20px;\n"
                                                              "}\n"
                                                              "\n"
                                                              "QPushButton:hover {\n"
                                                              "	background-color: rgb(133, 166, 177);\n"
                                                              "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                              "}\n"
                                                              "\n"
                                                              "QPushButton:pressed {\n"
                                                              "	background-color: rgb(133, 166, 177);\n"
                                                              "	font: 63 21pt \"Segoe UI Semibold\";\n"
                                                              "}")

                        self.ui.logo_notphoto_warning.hide()
                        self.ui.logo_notunique_warning.hide()

                    else:
                        self.ui.logo_notphoto_warning.raise_()
                        self.ui.logo_notphoto_warning.show()
                        self.ui.logo_notunique_warning.hide()
                        if delete_file != '':
                            self.ui.logo_browse_btn.setText('קיים לוגו, לחץ לשינוי')
                else:
                    self.ui.logo_notunique_warning.raise_()
                    self.ui.logo_notunique_warning.show()
                    self.ui.logo_notphoto_warning.hide()
                    if delete_file != '':
                        self.ui.logo_browse_btn.setText('קיים לוגו, לחץ לשינוי')

        elif 'reset' in btn_name:
            if delete_file != '':
                os.remove(folder_path + '/' + delete_file)  # delete the old image file
            excel_image = ''
            self.clear_all()
        else:
            self.clear_all()

            if delete_file != '':
                excel_image = folder_path + '/' + delete_file
                self.ui.logo_browse_btn.setText('קיים לוגו, לחץ לשינוי')

    def play_GIF(self):
        self.movie.start()

    def pause_GIF(self):
        self.movie.stop()

    def next_GIF(self):
        global gif_id
        gif_id += 1
        if gif_id == len(gif_dict) + 1:
            gif_id = 1

        self.movie = QMovie(gif_dict[gif_id])
        self.ui.GIF_label.setMovie(self.movie)
        self.movie.start()

    def stacked_homepage(self):
        self.ui.stacked_events_and_leads.setCurrentWidget(self.ui.home_page)

    ## EVENTS FUNCTIONS
    def stacked_events(self):
        self.ui.stacked_events_and_leads.setCurrentWidget(self.ui.events_page)

    def event_pages_and_buttons(self):
        global events_dict
        global event_pages
        global events_counter

        for page in range(0, event_pages):
            self.events_page_new = QWidget()
            self.events_page_new.setObjectName(u"events_page_" + str(page))
            self.ui.events_page_stacked.addWidget(self.events_page_new)
            self.ui.next_events_btn.raise_()

            temp_events = []

            if page == 0:
                events_counter = 0
                for temp_event in events[events_counter:events_counter + 9]:
                    temp_events.append(temp_event)

            elif page != 0:
                for temp_event in events[events_counter:events_counter + 9]:
                    temp_events.append(temp_event)

            row = 0

            event_alert_x = 0
            event_alert_y = 0
            EVENT_ALERT_WIDTH = 41
            EVENT_ALERT_HEIGHT = 41

            event_date_x = 0
            event_date_y = 0
            EVENT_DATE_WIDTH = 233
            EVENT_DATE_HEIGHT = 31

            event_button_x = 0
            event_button_y = 0
            EVENT_BUTTON_WIDTH = 291
            EVENT_BUTTON_HEIGHT = 71

            for event in temp_events:

                # COUNTDOWN CALCULATION
                event_date = re.findall(r"[\w']+", event[1])
                try:
                    if len(event_date[2]) == 2:
                        event_date[2] = int(event_date[2]) + 2000
                    time_to_event = (
                            date(int(event_date[2]), int(event_date[1]), int(event_date[0])) - date.today()).days
                    days_countdown = 'בעוד ' + str(time_to_event) + ' ימים'
                    new_date = str(event_date[0]) + '.' + str(event_date[1]) + '.' + str(event_date[2])
                    countdown = days_countdown + '  |  ' + new_date


                except:
                    countdown = emojize('לא הוזן תאריך ' + ':neutral_face:')

                if events_counter % 9 == 0:
                    events_counter += 1

                    event_alert_x += 270
                    event_alert_y += 20

                    event_date_x += 40
                    event_date_y += 100

                    event_button_x += 10
                    event_button_y += 40

                    self.event_alert = QPushButton(self.events_page_new)
                    self.event_alert.setObjectName(event[0] + '_alert_button')
                    self.event_alert.setGeometry(QRect(event_alert_x, event_alert_y,
                                                       EVENT_ALERT_WIDTH, EVENT_ALERT_HEIGHT))
                    self.event_alert.setFont(font101)
                    self.event_alert.setCursor(QCursor(Qt.PointingHandCursor))
                    self.event_alert.setMouseTracking(True)
                    self.event_alert.setFocusPolicy(Qt.ClickFocus)
                    self.event_alert.setStyleSheet(u"QPushButton { \n"
                                                   "	background-color: rgb(255,174,25);\n"
                                                   "background-color: rgb(226,0,57);\n"
                                                   "	color: rgb(255, 255, 255);\n"
                                                   "	font: 63 15pt \"Segoe UI Semilight\";\n"
                                                   "	border-radius: 20px;\n"
                                                   "}\n"
                                                   "\n"
                                                   "QPushButton:hover {\n"
                                                   "	font: 63 15pt \"Segoe UI Semibold\";\n"
                                                   "}\n"
                                                   "\n"
                                                   "QPushButton:pressed {\n"
                                                   "	background-color: rgb(229,156,22);\n"
                                                   "}")

                    mission_cnt = 0
                    records = DataBase.fetch_from_db(self, 'missions', 'all', event[0])
                    for record in records:
                        if record[3] == 'בתהליך':
                            mission_cnt += 1

                    self.event_alert.setText(str(mission_cnt))
                    self.event_alert.setWhatsThis('מספר המשימות הפתוחות')
                    self.event_alert.setToolTip('מספר המשימות הפתוחות')

                    self.event_date = QLabel(self.events_page_new)
                    self.event_date.setObjectName(event[0] + '_date_label')
                    self.event_date.setGeometry(QRect(event_date_x, event_date_y,
                                                      EVENT_DATE_WIDTH, EVENT_DATE_HEIGHT))
                    self.event_date.setStyleSheet(u"QLabel {\n"
                                                  "	background-color: rgb(189,233,216);\n"
                                                  "	color: rgb(255, 255, 255);\n"
                                                  "	color: rgb(0, 0, 0);\n"
                                                  "	\n"
                                                  "	font: 63 9pt \"Segoe UI Semibold\";\n"
                                                  "	border-radius: 15px;\n"
                                                  "}")
                    self.event_date.setFrameShape(QFrame.NoFrame)
                    self.event_date.setLineWidth(0)
                    self.event_date.setTextFormat(Qt.AutoText)
                    self.event_date.setScaledContents(False)
                    self.event_date.setAlignment(Qt.AlignCenter)
                    self.event_date.setWordWrap(False)
                    self.event_date.setText(countdown)

                    self.event_btn = QPushButton(self.events_page_new)
                    self.event_btn.setObjectName(event[0] + '_button')
                    self.event_btn.setGeometry(QRect(event_button_x, event_button_y,
                                                     EVENT_BUTTON_WIDTH, EVENT_BUTTON_HEIGHT))
                    self.event_btn.setFont(font101)
                    self.event_btn.setCursor(QCursor(Qt.PointingHandCursor))
                    self.event_btn.setMouseTracking(True)
                    self.event_btn.setFocusPolicy(Qt.ClickFocus)
                    self.event_btn.setStyleSheet(u"QPushButton {\n"
                                                 "	background-color: rgb(137,236,218);\n"
                                                 "	color: rgb(255, 255, 255);\n"
                                                 "	font: 63 15pt \"Segoe UI Semilight\";\n"
                                                 "	border-radius: 30px;\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:hover {\n"
                                                 "	font: 63 16pt \"Segoe UI Semibold\";\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:pressed {\n"
                                                 "	background-color: rgb(151,186,172);\n"
                                                 "}")
                    self.event_btn.setText(event[0])

                    # DICTIONARY-  {EVENT_NAME : BUTTON-ID}
                    startIndex = str(self.event_btn).find('(') + 1
                    endIndex = str(self.event_btn).find(',', startIndex + 1)
                    btn_id = str(self.event_btn)[startIndex:endIndex].split("x")[1]
                    events_dict[btn_id] = event[0]

                    self.event_btn.clicked.connect(self.get_event_btn_name)

                    self.event_alert.clicked.connect(self.get_event_alerts_btn_name)

                    self.event_btn.raise_()
                    self.event_date.raise_()
                    self.event_alert.raise_()

                    continue

                elif events_counter % 3 != 0:
                    event_alert_x += 330
                    event_alert_y += 0

                    event_date_x += 330
                    event_date_y += 0

                    event_button_x += 330
                    event_button_y += 0

                    self.event_alert = QPushButton(self.events_page_new)
                    self.event_alert.setObjectName(event[0] + '_alert_button')
                    self.event_alert.setGeometry(QRect(event_alert_x, event_alert_y,
                                                       EVENT_ALERT_WIDTH, EVENT_ALERT_HEIGHT))
                    self.event_alert.setFont(font101)
                    self.event_alert.setCursor(QCursor(Qt.PointingHandCursor))
                    self.event_alert.setMouseTracking(True)
                    self.event_alert.setFocusPolicy(Qt.ClickFocus)
                    self.event_alert.setStyleSheet(u"QPushButton { \n"
                                                   "	background-color: rgb(255,174,25);\n"
                                                   "background-color: rgb(226,0,57);\n"
                                                   "	color: rgb(255, 255, 255);\n"
                                                   "	font: 63 15pt \"Segoe UI Semilight\";\n"
                                                   "	border-radius: 20px;\n"
                                                   "}\n"
                                                   "\n"
                                                   "QPushButton:hover {\n"
                                                   "	font: 63 15pt \"Segoe UI Semibold\";\n"
                                                   "}\n"
                                                   "\n"
                                                   "QPushButton:pressed {\n"
                                                   "	background-color: rgb(229,156,22);\n"
                                                   "}")

                    mission_cnt = 0
                    records = DataBase.fetch_from_db(self, 'missions', 'all', event[0])
                    for record in records:
                        if record[3] == 'בתהליך':
                            mission_cnt += 1

                    self.event_alert.setText(str(mission_cnt))
                    self.event_alert.setWhatsThis('מספר המשימות הפתוחות')
                    self.event_alert.setToolTip('מספר המשימות הפתוחות')

                    self.event_date = QLabel(self.events_page_new)
                    self.event_date.setObjectName(event[0] + '_date_label')
                    self.event_date.setGeometry(QRect(event_date_x, event_date_y,
                                                      EVENT_DATE_WIDTH, EVENT_DATE_HEIGHT))
                    self.event_date.setStyleSheet(u"QLabel {\n"
                                                  "	background-color: rgb(189,233,216);\n"
                                                  "	color: rgb(255, 255, 255);\n"
                                                  "	color: rgb(0, 0, 0);\n"
                                                  "	\n"
                                                  "	font: 63 9pt \"Segoe UI Semibold\";\n"
                                                  "	border-radius: 15px;\n"
                                                  "}")
                    self.event_date.setFrameShape(QFrame.NoFrame)
                    self.event_date.setLineWidth(0)
                    self.event_date.setTextFormat(Qt.AutoText)
                    self.event_date.setScaledContents(False)
                    self.event_date.setAlignment(Qt.AlignCenter)
                    self.event_date.setWordWrap(False)
                    self.event_date.setText(countdown)

                    self.event_btn = QPushButton(self.events_page_new)
                    self.event_btn.setObjectName(event[0] + '_button')
                    self.event_btn.setGeometry(QRect(event_button_x, event_button_y,
                                                     EVENT_BUTTON_WIDTH, EVENT_BUTTON_HEIGHT))
                    self.event_btn.setFont(font101)
                    self.event_btn.setCursor(QCursor(Qt.PointingHandCursor))
                    self.event_btn.setMouseTracking(True)
                    self.event_btn.setFocusPolicy(Qt.ClickFocus)
                    self.event_btn.setStyleSheet(u"QPushButton {\n"
                                                 "	background-color: rgb(137,236,218);\n"
                                                 "	color: rgb(255, 255, 255);\n"
                                                 "	font: 63 15pt \"Segoe UI Semilight\";\n"
                                                 "	border-radius: 30px;\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:hover {\n"
                                                 "	font: 63 16pt \"Segoe UI Semibold\";\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:pressed {\n"
                                                 "	background-color: rgb(151,186,172);\n"
                                                 "}")
                    self.event_btn.setText(event[0])

                    # DICTIONARY-  {EVENT_NAME : BUTTON-ID}
                    startIndex = str(self.event_btn).find('(') + 1
                    endIndex = str(self.event_btn).find(',', startIndex + 1)
                    btn_id = str(self.event_btn)[startIndex:endIndex].split("x")[1]
                    events_dict[btn_id] = event[0]

                    self.event_btn.clicked.connect(self.get_event_btn_name)
                    self.event_alert.clicked.connect(self.get_event_alerts_btn_name)

                    self.event_btn.raise_()
                    self.event_date.raise_()
                    self.event_alert.raise_()

                    events_counter += 1
                    continue

                elif events_counter % 3 == 0 and events_counter % 9 != 0:
                    row += 1
                    event_alert_x = 270
                    event_alert_y = 20 + row * 130

                    event_date_x = 40
                    event_date_y = 100 + row * 130

                    event_button_x = 10
                    event_button_y = 40 + row * 130

                    self.event_alert = QPushButton(self.events_page_new)
                    self.event_alert.setObjectName(event[0] + '_alert_button')
                    self.event_alert.setGeometry(QRect(event_alert_x, event_alert_y,
                                                       EVENT_ALERT_WIDTH, EVENT_ALERT_HEIGHT))
                    self.event_alert.setFont(font101)
                    self.event_alert.setCursor(QCursor(Qt.PointingHandCursor))
                    self.event_alert.setMouseTracking(True)
                    self.event_alert.setFocusPolicy(Qt.ClickFocus)
                    self.event_alert.setStyleSheet(u"QPushButton { \n"
                                                   "	background-color: rgb(255,174,25);\n"
                                                   "background-color: rgb(226,0,57);\n"
                                                   "	color: rgb(255, 255, 255);\n"
                                                   "	font: 63 15pt \"Segoe UI Semilight\";\n"
                                                   "	border-radius: 20px;\n"
                                                   "}\n"
                                                   "\n"
                                                   "QPushButton:hover {\n"
                                                   "	font: 63 15pt \"Segoe UI Semibold\";\n"
                                                   "}\n"
                                                   "\n"
                                                   "QPushButton:pressed {\n"
                                                   "	background-color: rgb(229,156,22);\n"
                                                   "}")
                    mission_cnt = 0
                    records = DataBase.fetch_from_db(self, 'missions', 'all', event[0])
                    for record in records:
                        if record[3] == 'בתהליך':
                            mission_cnt += 1

                    self.event_alert.setText(str(mission_cnt))
                    self.event_alert.setWhatsThis('מספר המשימות הפתוחות')
                    self.event_alert.setToolTip('מספר המשימות הפתוחות')

                    self.event_date = QLabel(self.events_page_new)
                    self.event_date.setObjectName(event[0] + '_date_label')
                    self.event_date.setGeometry(QRect(event_date_x, event_date_y,
                                                      EVENT_DATE_WIDTH, EVENT_DATE_HEIGHT))
                    self.event_date.setStyleSheet(u"QLabel {\n"
                                                  "	background-color: rgb(189,233,216);\n"
                                                  "	color: rgb(255, 255, 255);\n"
                                                  "	color: rgb(0, 0, 0);\n"
                                                  "	\n"
                                                  "	font: 63 9pt \"Segoe UI Semibold\";\n"
                                                  "	border-radius: 15px;\n"
                                                  "}")
                    self.event_date.setFrameShape(QFrame.NoFrame)
                    self.event_date.setLineWidth(0)
                    self.event_date.setTextFormat(Qt.AutoText)
                    self.event_date.setScaledContents(False)
                    self.event_date.setAlignment(Qt.AlignCenter)
                    self.event_date.setWordWrap(False)
                    self.event_date.setText(countdown)

                    self.event_btn = QPushButton(self.events_page_new)
                    self.event_btn.setObjectName(event[0] + '_button')
                    self.event_btn.setGeometry(QRect(event_button_x, event_button_y,
                                                     EVENT_BUTTON_WIDTH, EVENT_BUTTON_HEIGHT))
                    self.event_btn.setFont(font101)
                    self.event_btn.setCursor(QCursor(Qt.PointingHandCursor))
                    self.event_btn.setMouseTracking(True)
                    self.event_btn.setFocusPolicy(Qt.ClickFocus)
                    self.event_btn.setStyleSheet(u"QPushButton {\n"
                                                 "	background-color: rgb(137,236,218);\n"
                                                 "	color: rgb(255, 255, 255);\n"
                                                 "	font: 63 15pt \"Segoe UI Semilight\";\n"
                                                 "	border-radius: 30px;\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:hover {\n"
                                                 "	font: 63 16pt \"Segoe UI Semibold\";\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:pressed {\n"
                                                 "	background-color: rgb(151,186,172);\n"
                                                 "}")
                    self.event_btn.setText(event[0])

                    # DICTIONARY-  {EVENT_NAME : BUTTON-ID}
                    startIndex = str(self.event_btn).find('(') + 1
                    endIndex = str(self.event_btn).find(',', startIndex + 1)
                    btn_id = str(self.event_btn)[startIndex:endIndex].split("x")[1]
                    events_dict[btn_id] = event[0]

                    self.event_btn.clicked.connect(self.get_event_btn_name)
                    self.event_alert.clicked.connect(self.get_event_alerts_btn_name)

                    self.event_btn.raise_()
                    self.event_date.raise_()
                    self.event_alert.raise_()

                    events_counter += 1
                    continue

                else:
                    break

    def next_events_page_btn(self):
        self.ui.events_page_stacked.setCurrentIndex(self.ui.events_page_stacked.currentIndex() + 1)
        self.ui.events_page_number_label.setText(str(self.ui.events_page_stacked.currentIndex()))
        if self.ui.events_page_stacked.currentIndex() == (event_pages - 1):
            self.ui.next_events_btn.hide()
        self.ui.previous_events_btn.show()
        self.ui.previous_events_btn.raise_()

    def previous_events_page_btn(self):
        self.ui.events_page_stacked.setCurrentIndex((self.ui.events_page_stacked.currentIndex() - 1) % event_pages)
        self.ui.events_page_number_label.setText(str(self.ui.events_page_stacked.currentIndex()))
        if self.ui.events_page_stacked.currentIndex() == 0:
            self.ui.previous_events_btn.hide()
        self.ui.next_events_btn.show()

    def get_event_btn_name(self):
        event_btn_name = self.sender()

        startIndex = str(event_btn_name).find('(') + 1
        endIndex = str(event_btn_name).find(',', startIndex + 1)
        btn_id = str(event_btn_name)[startIndex:endIndex].split("x")[1]

        self.event_window.pop_window(events_dict[btn_id])

    def get_event_alerts_btn_name(self):
        event_alerts_btn_name = self.sender()
        self.event_window.pop_alerts(event_alerts_btn_name)

    ## LEADS FUNCTIONS
    def stacked_leads(self):
        self.ui.stacked_events_and_leads.setCurrentWidget(self.ui.leads_page)

    def lead_pages_and_buttons(self):
        global leads_dict
        global lead_pages
        global leads_counter

        for page in range(0, lead_pages):
            self.leads_page_new = QWidget()
            self.leads_page_new.setObjectName(u"leads_page_" + str(page))
            self.ui.leads_page_stacked.addWidget(self.leads_page_new)
            self.ui.next_leads_btn.raise_()

            temp_leads = []

            if page == 0:
                leads_counter = 0
                for temp_lead in leads[leads_counter:leads_counter + 9]:
                    temp_leads.append(temp_lead)

            elif page != 0:
                for temp_lead in leads[leads_counter:leads_counter + 9]:
                    temp_leads.append(temp_lead)

            row = 0

            lead_alert_x = 0
            lead_alert_y = 0
            LEAD_ALERT_WIDTH = 41
            LEAD_ALERT_HEIGHT = 41

            lead_date_x = 0
            lead_date_y = 0
            LEAD_DATE_WIDTH = 233
            LEAD_DATE_HEIGHT = 31

            lead_button_x = 0
            lead_button_y = 0
            LEAD_BUTTON_WIDTH = 291
            LEAD_BUTTON_HEIGHT = 71

            for lead in temp_leads:

                # COUNTDOWN CALCULATION
                lead_date = re.findall(r"[\w']+", lead[1])
                try:
                    if len(lead_date[2]) == 2:
                        lead_date[2] = int(lead_date[2]) + 2000
                    time_to_lead = (
                            date(int(lead_date[2]), int(lead_date[1]), int(lead_date[0])) - date.today()).days
                    days_countdown = 'בעוד ' + str(time_to_lead) + ' ימים'
                    new_date = lead_date[0] + '.' + lead_date[1] + '.' + lead_date[2]
                    countdown = days_countdown + '  |  ' + new_date

                except:
                    countdown = emojize('לא הוזן תאריך ' + ':neutral_face:')

                if leads_counter % 9 == 0:
                    leads_counter += 1

                    lead_alert_x += 270
                    lead_alert_y += 20

                    lead_date_x += 40
                    lead_date_y += 100

                    lead_button_x += 10
                    lead_button_y += 40

                    self.lead_alert = QPushButton(self.leads_page_new)
                    self.lead_alert.setObjectName(lead[0] + '_alert_button')
                    self.lead_alert.setGeometry(QRect(lead_alert_x, lead_alert_y,
                                                      LEAD_ALERT_WIDTH, LEAD_ALERT_HEIGHT))
                    self.lead_alert.setFont(font100)
                    self.lead_alert.setCursor(QCursor(Qt.PointingHandCursor))
                    self.lead_alert.setMouseTracking(True)
                    self.lead_alert.setFocusPolicy(Qt.ClickFocus)
                    self.lead_alert.setStyleSheet(u"QPushButton { \n"
                                                  "    background-color: rgb(255,174,25);\n"
                                                  "    background-color: rgb(226,0,57);\n"
                                                  "    color: rgb(255, 255, 255);\n"
                                                  "    font: 63 15pt \"Segoe UI Semilight\";\n"
                                                  "    border-radius: 20px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "QPushButton:hover {\n"
                                                  "     font: 63 15pt \"Segoe UI Semibold\";\n"
                                                  "}\n"
                                                  "\n"
                                                  "QPushButton:pressed {\n"
                                                  "     background-color: rgb(229,156,22);\n"
                                                  "}")
                    self.lead_alert.setText('5')

                    self.lead_date = QLabel(self.leads_page_new)
                    self.lead_date.setObjectName(lead[0] + '_date_label')
                    self.lead_date.setGeometry(QRect(lead_date_x, lead_date_y,
                                                     LEAD_DATE_WIDTH, LEAD_DATE_HEIGHT))
                    self.lead_date.setStyleSheet(u"QLabel {\n"
                                                 "  background-color: rgb(215,227,246);\n"
                                                 "	color: rgb(0, 0, 0);	\n"
                                                 "	font: 63 9pt \"Segoe UI Semibold\";\n"
                                                 "	border-radius: 15px;\n"
                                                 "}")
                    self.lead_date.setFrameShape(QFrame.NoFrame)
                    self.lead_date.setLineWidth(0)
                    self.lead_date.setTextFormat(Qt.AutoText)
                    self.lead_date.setScaledContents(False)
                    self.lead_date.setAlignment(Qt.AlignCenter)
                    self.lead_date.setWordWrap(False)
                    self.lead_date.setText(countdown)

                    self.lead_btn = QPushButton(self.leads_page_new)
                    self.lead_btn.setObjectName(lead[0] + '_button')
                    self.lead_btn.setGeometry(QRect(lead_button_x, lead_button_y,
                                                    LEAD_BUTTON_WIDTH, LEAD_BUTTON_HEIGHT))
                    self.lead_btn.setFont(font100)
                    self.lead_btn.setCursor(QCursor(Qt.PointingHandCursor))
                    self.lead_btn.setMouseTracking(True)
                    self.lead_btn.setFocusPolicy(Qt.ClickFocus)
                    self.lead_btn.setStyleSheet(u"QPushButton {\n"
                                                "	background-color: rgb(196,213,242);\n"
                                                "	color: rgb(255, 255, 255);\n"
                                                "	font: 63 15pt \"Segoe UI Semilight\";\n"
                                                "	border-radius: 30px;\n"
                                                "}\n"
                                                "\n"
                                                "QPushButton:hover {\n"
                                                "	font: 63 16pt \"Segoe UI Semibold\";\n"
                                                "}\n"
                                                "\n"
                                                "QPushButton:pressed {\n"
                                                "   background-color: rgb(196,197,242);\n"
                                                "}")
                    self.lead_btn.setText(lead[0])

                    # DICTIONARY-  {LEAD_NAME : BUTTON-ID}
                    startIndex = str(self.lead_btn).find('(') + 1
                    endIndex = str(self.lead_btn).find(',', startIndex + 1)
                    btn_id = str(self.lead_btn)[startIndex:endIndex].split("x")[1]
                    leads_dict[btn_id] = lead[0]

                    self.lead_btn.clicked.connect(self.get_lead_btn_name)
                    self.lead_alert.clicked.connect(self.get_lead_alerts_btn_name)

                    self.lead_btn.raise_()
                    self.lead_date.raise_()
                    self.lead_alert.raise_()

                    continue

                elif leads_counter % 3 != 0:
                    lead_alert_x += 330
                    lead_alert_y += 0

                    lead_date_x += 330
                    lead_date_y += 0

                    lead_button_x += 330
                    lead_button_y += 0

                    self.lead_alert = QPushButton(self.leads_page_new)
                    self.lead_alert.setObjectName(lead[0] + '_alert_button')
                    self.lead_alert.setGeometry(QRect(lead_alert_x, lead_alert_y,
                                                      LEAD_ALERT_WIDTH, LEAD_ALERT_HEIGHT))
                    self.lead_alert.setFont(font100)
                    self.lead_alert.setCursor(QCursor(Qt.PointingHandCursor))
                    self.lead_alert.setMouseTracking(True)
                    self.lead_alert.setFocusPolicy(Qt.ClickFocus)
                    self.lead_alert.setStyleSheet(u"QPushButton { \n"
                                                  "    background-color: rgb(255,174,25);\n"
                                                  "    background-color: rgb(226,0,57);\n"
                                                  "    color: rgb(255, 255, 255);\n"
                                                  "    font: 63 15pt \"Segoe UI Semilight\";\n"
                                                  "    border-radius: 20px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "QPushButton:hover {\n"
                                                  "     font: 63 15pt \"Segoe UI Semibold\";\n"
                                                  "}\n"
                                                  "\n"
                                                  "QPushButton:pressed {\n"
                                                  "     background-color: rgb(229,156,22);\n"
                                                  "}")
                    self.lead_alert.setText('10')

                    self.lead_date = QLabel(self.leads_page_new)
                    self.lead_date.setObjectName(lead[0] + '_date_label')
                    self.lead_date.setGeometry(QRect(lead_date_x, lead_date_y,
                                                     LEAD_DATE_WIDTH, LEAD_DATE_HEIGHT))
                    self.lead_date.setStyleSheet(u"QLabel {\n"
                                                 "  background-color: rgb(215,227,246);\n"
                                                 "	color: rgb(0, 0, 0);	\n"
                                                 "	font: 63 9pt \"Segoe UI Semibold\";\n"
                                                 "	border-radius: 15px;\n"
                                                 "}")
                    self.lead_date.setFrameShape(QFrame.NoFrame)
                    self.lead_date.setLineWidth(0)
                    self.lead_date.setTextFormat(Qt.AutoText)
                    self.lead_date.setScaledContents(False)
                    self.lead_date.setAlignment(Qt.AlignCenter)
                    self.lead_date.setWordWrap(False)
                    self.lead_date.setText(countdown)

                    self.lead_btn = QPushButton(self.leads_page_new)
                    self.lead_btn.setObjectName(lead[0] + '_button')
                    self.lead_btn.setGeometry(QRect(lead_button_x, lead_button_y,
                                                    LEAD_BUTTON_WIDTH, LEAD_BUTTON_HEIGHT))
                    self.lead_btn.setFont(font100)
                    self.lead_btn.setCursor(QCursor(Qt.PointingHandCursor))
                    self.lead_btn.setMouseTracking(True)
                    self.lead_btn.setFocusPolicy(Qt.ClickFocus)
                    self.lead_btn.setStyleSheet(u"QPushButton {\n"
                                                "	background-color: rgb(196,213,242);\n"
                                                "	color: rgb(255, 255, 255);\n"
                                                "	font: 63 15pt \"Segoe UI Semilight\";\n"
                                                "	border-radius: 30px;\n"
                                                "}\n"
                                                "\n"
                                                "QPushButton:hover {\n"
                                                "	font: 63 16pt \"Segoe UI Semibold\";\n"
                                                "}\n"
                                                "\n"
                                                "QPushButton:pressed {\n"
                                                "   background-color: rgb(196,197,242);\n"
                                                "}")
                    self.lead_btn.setText(lead[0])

                    # DICTIONARY-  {LEAD_NAME : BUTTON-ID}
                    startIndex = str(self.lead_btn).find('(') + 1
                    endIndex = str(self.lead_btn).find(',', startIndex + 1)
                    btn_id = str(self.lead_btn)[startIndex:endIndex].split("x")[1]
                    leads_dict[btn_id] = lead[0]

                    self.lead_btn.clicked.connect(self.get_lead_btn_name)
                    self.lead_alert.clicked.connect(self.get_lead_alerts_btn_name)

                    self.lead_btn.raise_()
                    self.lead_date.raise_()
                    self.lead_alert.raise_()

                    leads_counter += 1
                    continue

                elif leads_counter % 3 == 0 and leads_counter % 9 != 0:
                    row += 1
                    lead_alert_x = 270
                    lead_alert_y = 20 + row * 130

                    lead_date_x = 60
                    lead_date_y = 100 + row * 130

                    lead_button_x = 10
                    lead_button_y = 40 + row * 130

                    self.lead_alert = QPushButton(self.leads_page_new)
                    self.lead_alert.setObjectName(lead[0] + '_alert_button')
                    self.lead_alert.setGeometry(QRect(lead_alert_x, lead_alert_y,
                                                      LEAD_ALERT_WIDTH, LEAD_ALERT_HEIGHT))
                    self.lead_alert.setFont(font100)
                    self.lead_alert.setCursor(QCursor(Qt.PointingHandCursor))
                    self.lead_alert.setMouseTracking(True)
                    self.lead_alert.setFocusPolicy(Qt.ClickFocus)
                    self.lead_alert.setStyleSheet(u"QPushButton { \n"
                                                  "    background-color: rgb(255,174,25);\n"
                                                  "    background-color: rgb(226,0,57);\n"
                                                  "    color: rgb(255, 255, 255);\n"
                                                  "    font: 63 15pt \"Segoe UI Semilight\";\n"
                                                  "    border-radius: 20px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "QPushButton:hover {\n"
                                                  "     font: 63 15pt \"Segoe UI Semibold\";\n"
                                                  "}\n"
                                                  "\n"
                                                  "QPushButton:pressed {\n"
                                                  "     background-color: rgb(229,156,22);\n"
                                                  "}")
                    self.lead_alert.setText('10')

                    self.lead_date = QLabel(self.leads_page_new)
                    self.lead_date.setObjectName(lead[0] + '_date_label')
                    self.lead_date.setGeometry(QRect(lead_date_x, lead_date_y,
                                                     LEAD_DATE_WIDTH, LEAD_DATE_HEIGHT))
                    self.lead_date.setStyleSheet(u"QLabel {\n"
                                                 "  background-color: rgb(215,227,246);\n"
                                                 "	color: rgb(0, 0, 0);	\n"
                                                 "	font: 63 9pt \"Segoe UI Semibold\";\n"
                                                 "	border-radius: 15px;\n"
                                                 "}")
                    self.lead_date.setFrameShape(QFrame.NoFrame)
                    self.lead_date.setLineWidth(0)
                    self.lead_date.setTextFormat(Qt.AutoText)
                    self.lead_date.setScaledContents(False)
                    self.lead_date.setAlignment(Qt.AlignCenter)
                    self.lead_date.setWordWrap(False)
                    self.lead_date.setText(countdown)

                    self.lead_btn = QPushButton(self.leads_page_new)
                    self.lead_btn.setObjectName(lead[0] + '_button')
                    self.lead_btn.setGeometry(QRect(lead_button_x, lead_button_y,
                                                    LEAD_BUTTON_WIDTH, LEAD_BUTTON_HEIGHT))
                    self.lead_btn.setFont(font100)
                    self.lead_btn.setCursor(QCursor(Qt.PointingHandCursor))
                    self.lead_btn.setMouseTracking(True)
                    self.lead_btn.setFocusPolicy(Qt.ClickFocus)
                    self.lead_btn.setStyleSheet(u"QPushButton {\n"
                                                "	background-color: rgb(196,213,242);\n"
                                                "	color: rgb(255, 255, 255);\n"
                                                "	font: 63 15pt \"Segoe UI Semilight\";\n"
                                                "	border-radius: 30px;\n"
                                                "}\n"
                                                "\n"
                                                "QPushButton:hover {\n"
                                                "	font: 63 16pt \"Segoe UI Semibold\";\n"
                                                "}\n"
                                                "\n"
                                                "QPushButton:pressed {\n"
                                                "   background-color: rgb(196,197,242);\n"
                                                "}")
                    self.lead_btn.setText(lead[0])

                    # DICTIONARY-  {LEAD_NAME : BUTTON-ID}
                    startIndex = str(self.lead_btn).find('(') + 1
                    endIndex = str(self.lead_btn).find(',', startIndex + 1)
                    btn_id = str(self.lead_btn)[startIndex:endIndex].split("x")[1]
                    leads_dict[btn_id] = lead[0]

                    self.lead_btn.clicked.connect(self.get_lead_btn_name)
                    self.lead_alert.clicked.connect(self.get_lead_alerts_btn_name)

                    self.lead_btn.raise_()
                    self.lead_date.raise_()
                    self.lead_alert.raise_()

                    leads_counter += 1
                    continue

                else:
                    break

    def next_leads_page_btn(self):
        self.ui.leads_page_stacked.setCurrentIndex(self.ui.leads_page_stacked.currentIndex() + 1)
        self.ui.leads_page_number_label.setText(str(self.ui.leads_page_stacked.currentIndex()))
        if self.ui.leads_page_stacked.currentIndex() == (lead_pages - 1):
            self.ui.next_leads_btn.hide()
        self.ui.previous_leads_btn.show()
        self.ui.previous_leads_btn.raise_()

    def previous_leads_page_btn(self):
        self.ui.leads_page_stacked.setCurrentIndex((self.ui.leads_page_stacked.currentIndex() - 1) % lead_pages)
        self.ui.leads_page_number_label.setText(str(self.ui.leads_page_stacked.currentIndex()))
        if self.ui.leads_page_stacked.currentIndex() == 0:
            self.ui.previous_leads_btn.hide()
        self.ui.next_leads_btn.show()

    def get_lead_btn_name(self):
        lead_btn_name = self.sender()

        startIndex = str(lead_btn_name).find('(') + 1
        endIndex = str(lead_btn_name).find(',', startIndex + 1)
        btn_id = str(lead_btn_name)[startIndex:endIndex].split("x")[1]

        self.lead_window.pop_window(leads_dict[btn_id])

    def get_lead_alerts_btn_name(self):
        lead_alerts_btn_name = self.sender()
        self.lead_window.pop_alerts(lead_alerts_btn_name)

    def btn_to_web(self, web):
        if web == "whatsapp":
            webbrowser.open_new("https://web.whatsapp.com/")
        elif web == "fb":
            webbrowser.open_new("https://facebook.com/")
        elif web == "instagram":
            webbrowser.open_new("https://instagram.com//")
        elif web == "calender":
            webbrowser.open_new("https://calendar.google.com/calendar/u/0/r")
        elif web == "gmail":
            webbrowser.open_new("https://mail.google.com//")

    def clear_all(self):
        self.ui.logo_browse_btn.hide()
        self.ui.reset_logo_btn.hide()
        self.ui.logo_notphoto_warning.hide()
        self.ui.logo_notunique_warning.hide()

        self.ui.line.hide()

        self.ui.backup_Setup_btn.hide()
        self.ui.restore_backup_btn.hide()
        self.ui.restart_backup_warning.hide()
        self.ui.restart_for_backup_btn.hide()
        self.ui.backup_freq_label.hide()
        self.ui.backup_options_combobox.hide()

        self.ui.nextGIF_btn.hide()
        self.ui.playGIF_btn.hide()
        self.ui.pauseGIF_btn.hide()

        self.ui.logo_browse_btn.setText('בחר לוגו עבור דו"חות')
        self.ui.logo_browse_btn.setStyleSheet(u"QPushButton {\n"
                                              "	background-color: rgb(170,170,170);	\n"
                                              "	color: rgb(255, 255, 255);\n"
                                              "	border: 1px solid #090206;\n"
                                              "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                              "	border-radius: 20px;\n"
                                              "}\n"
                                              "\n"
                                              "QPushButton:hover {\n"
                                              "	background-color: rgb(133, 166, 177);\n"
                                              "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                              "}\n"
                                              "\n"
                                              "QPushButton:pressed {\n"
                                              "	background-color: rgb(133, 166, 177);\n"
                                              "	font: 63 21pt \"Segoe UI Semibold\";\n"
                                              "}")


# CUSTOMERS WINDOW
class CustomersWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_CustomersWindow()
        self.ui.setupUi(self)
        self.setWindowIcon(QtGui.QIcon('Pheonix.ico'))

        # BUTTONS
        self.ui.add_client_btn.clicked.connect(self.add_customer)
        self.ui.edit_client_btn.clicked.connect(self.edit_customer)
        self.ui.update_client_btn.clicked.connect(self.update_customer)

        self.ui.single_select_delete_yes_btn.clicked.connect(self.remove_customer)
        self.ui.single_select_delete_no_btn.clicked.connect(self.cancel_remove)
        self.ui.delete_selection_client_btn.clicked.connect(self.check_selection)
        self.ui.all_select_delete_yes_btn.clicked.connect(self.remove_all_customers)
        self.ui.all_select_delete_no_btn.clicked.connect(self.cancel_remove)
        self.ui.delete_all_clients_btn.clicked.connect(self.check_selection)

        self.ui.clear_client_fields_btn.clicked.connect(self.clear_all)
        self.ui.customers_export_excel_btn.clicked.connect(self.export_to_excel)
        self.ui.lead_check.clicked.connect(self.lead_check_btn)

        # LABELS
        self.ui.delete_selection_client_btn.enterEvent = lambda e: self.ui.delete_selection_warning.show()
        self.ui.delete_selection_client_btn.leaveEvent = lambda e: self.ui.delete_selection_warning.hide()
        self.ui.delete_all_clients_btn.enterEvent = lambda e: self.ui.delete_all_warning.show()
        self.ui.delete_all_clients_btn.leaveEvent = lambda e: self.ui.delete_all_warning.hide()

    def lead_check_btn(self):
        if self.ui.lead_check.isChecked():
            self.ui.editing_state_label.setStyleSheet(u"font: 16pt \"Segoe UI Semibold\";\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "background-color: rgb(196,213,242);\n"
                                                      "border-radius: 20px;\n"
                                                      "")
            self.ui.editing_state_label.setText('מצב הוספת ליד')

        else:
            self.ui.editing_state_label.setStyleSheet(u"font: 16pt \"Segoe UI Semibold\";\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "background-color: rgb(137,236,218);\n"
                                                      "border-radius: 20px;\n"
                                                      "")
            self.ui.editing_state_label.setText('מצב הוספת אירוע')

    def export_to_excel(self):
        self.ui.main.hide()
        btn_clicked = self.sender()
        try:
            create_excel(btn_clicked)
        except:
            self.ui.excel_warning.show()

    def tree_setup(self):
        # TREEWIDGET VIEW
        self.ui.customers_tablewidget.setColumnWidth(0, 175)  # event_name
        self.ui.customers_tablewidget.setColumnWidth(1, 80)  # event_date
        self.ui.customers_tablewidget.setColumnWidth(2, 110)  # contact_name
        self.ui.customers_tablewidget.setColumnWidth(3, 180)  # contact_address
        self.ui.customers_tablewidget.setColumnWidth(4, 95)  # contact_phone
        self.ui.customers_tablewidget.setColumnWidth(5, 200)  # contact_email
        self.ui.customers_tablewidget.setColumnWidth(6, 60)  # contract
        self.ui.customers_tablewidget.setColumnWidth(7, 55)  # rate
        self.ui.customers_tablewidget.setColumnWidth(8, 113)  # location

        self.load_customers_tree()

    def load_customers_tree(self):
        row = 0
        event_color = {'כן': QtGui.QColor(196, 197, 242),
                       'לא': QtGui.QColor(137, 236, 218)
                       }

        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("SELECT *,oid FROM customers")
        customers = cursor.fetchall()
        cursor.close()
        db_connect.close()

        self.ui.customers_tablewidget.setRowCount(len(customers))

        for customer in customers:
            event_name, event_date, contact_name, contact_address, contact_phone, \
            contact_email, contract, rate, location, location_address, place, lead, uid = customer  ## set up variables from tuple

            self.ui.customers_tablewidget.setItem(row, 0,
                                                  QtWidgets.QTableWidgetItem(event_name))
            self.ui.customers_tablewidget.setItem(row, 1,
                                                  QtWidgets.QTableWidgetItem(event_date))
            self.ui.customers_tablewidget.setItem(row, 2,
                                                  QtWidgets.QTableWidgetItem(contact_name))
            self.ui.customers_tablewidget.setItem(row, 3,
                                                  QtWidgets.QTableWidgetItem(contact_address))
            self.ui.customers_tablewidget.setItem(row, 4,
                                                  QtWidgets.QTableWidgetItem(contact_phone))
            self.ui.customers_tablewidget.setItem(row, 5,
                                                  QtWidgets.QTableWidgetItem(contact_email))
            self.ui.customers_tablewidget.setItem(row, 6,
                                                  QtWidgets.QTableWidgetItem(contract))
            self.ui.customers_tablewidget.setItem(row, 7,
                                                  QtWidgets.QTableWidgetItem(make_amount_readable(rate)))
            self.ui.customers_tablewidget.setItem(row, 8,
                                                  QtWidgets.QTableWidgetItem(location))

            self.ui.customers_tablewidget.item(row, 0).setBackground(event_color[lead])
            row += 1

        self.ui.customers_tablewidget.resizeRowsToContents()

        if customers:
            self.ui.edit_client_btn.show()
        else:
            self.ui.edit_client_btn.hide()
        self.clear_all()

    def add_customer(self):
        place = self.ui.event_place_combobox.currentText()
        if self.ui.contact1_name_input.toPlainText() != '' and self.ui.event_name_input.toPlainText() != '':

            event_date = make_date_standard(self.ui.event_date_input.toPlainText())

            # CheckBox for contract
            if self.ui.contract_checkbox.isChecked():
                contract = 'קיים'
            else:
                contract = 'לא קיים'
            if self.ui.lead_check.isChecked():
                lead = 'כן'
            else:
                lead = 'לא'

            customer_1 = {'event_name': self.ui.event_name_input.toPlainText(),
                          'event_date': event_date,
                          'contact_name': self.ui.contact1_name_input.toPlainText(),
                          'contact_address': self.ui.contact1_address_input.toPlainText(),
                          'contact_phone': self.ui.contact1_phone_input.toPlainText(),
                          'contact_email': self.ui.contact1_email_input.toPlainText(),
                          'contract': contract,
                          'rate': self.ui.rate_input.toPlainText(),
                          'location': self.ui.event_location_input.toPlainText(),
                          'location_address': self.ui.event_location_address_input.toPlainText(),
                          'place': place,
                          'lead': lead}

            # add the places DB if place does not exist in it
            places = DataBase.fetch_from_db(self, 'places')
            if place not in places:
                DataBase.insert_into_db(self, 'places', place)

            DataBase.insert_into_db(self, 'customers', customer_1)

            if self.ui.contact2_name_input.toPlainText() != '':
                event_date = make_date_standard(self.ui.event_date_input.toPlainText())

                customer_2 = {'event_name': self.ui.event_name_input.toPlainText(),
                              'event_date': event_date,
                              'contact_name': self.ui.contact2_name_input.toPlainText(),
                              'contact_address': self.ui.contact2_address_input.toPlainText(),
                              'contact_phone': self.ui.contact2_phone_input.toPlainText(),
                              'contact_email': self.ui.contact2_email_input.toPlainText(),
                              'contract': contract,
                              'rate': self.ui.rate_input.toPlainText(),
                              'location': self.ui.event_location_input.toPlainText(),
                              'location_address': self.ui.event_location_address_input.toPlainText(),
                              'place': place,
                              'lead': lead}

                DataBase.insert_into_db(self, 'customers', customer_2)

            self.load_customers_tree()

        elif self.ui.event_name_input.toPlainText() == '' and self.ui.contact1_name_input.toPlainText() != '':
            self.ui.event_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                   "border: 3px solid #ef4040;\n"
                                                   "background-color: rgb(255,255,255);\n"
                                                   "border-radius: 20px;")
            self.ui.event_mendatory_label.show()
            self.ui.event_mendatory_label.raise_()

        elif self.ui.contact1_name_input.toPlainText() == '' and self.ui.event_name_input.toPlainText() != '':
            self.ui.contact1_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                      "border: 3px solid #ef4040;\n"
                                                      "background-color: rgb(255,255,255);\n"
                                                      "border-radius: 20px;")
            self.ui.contact_mendatory_label.show()
            self.ui.contact_mendatory_label.raise_()

        else:
            self.ui.event_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                   "border: 3px solid #ef4040;\n"
                                                   "background-color: rgb(255,255,255);\n"
                                                   "border-radius: 20px;")

            self.ui.contact1_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                      "border: 3px solid #ef4040;\n"
                                                      "background-color: rgb(255,255,255);\n"
                                                      "border-radius: 20px;")
            self.ui.contact_mendatory_label.show()
            self.ui.event_mendatory_label.show()
            self.ui.event_mendatory_label.raise_()
            self.ui.contact_mendatory_label.raise_()

    def edit_customer(self):
        global customer_oid
        self.ui.clear_client_fields_btn.setEnabled(False)
        self.ui.add_client_btn.setEnabled(False)
        self.ui.delete_all_clients_btn.setEnabled(False)
        self.ui.delete_selection_client_btn.setEnabled(False)
        self.ui.contact_mendatory_label.hide()
        self.ui.event_mendatory_label.hide()

        try:
            row = self.ui.customers_tablewidget.currentRow()
            column = 0
            event_name = self.ui.customers_tablewidget.item(row, column).text()
            contact_name = self.ui.customers_tablewidget.item(row, column + 2).text()

            record = DataBase.fetch_from_db(self, 'customers', 'one', event_name, contact_name)

            event_name, event_date, contact_name, contact_address, contact_phone, contact_email, contract,\
            rate, location, location_address, place, lead, customer_oid = record[0]  ## set up variables from tuple

            self.ui.event_name_input.setText(event_name)
            self.ui.event_date_input.setText(event_date)
            self.ui.contact1_name_input.setText(contact_name)
            self.ui.contact1_address_input.setText(contact_address)
            self.ui.contact1_phone_input.setText(contact_phone)
            self.ui.contact1_email_input.setText(contact_email)

            if contract == 'קיים':
                self.ui.contract_checkbox.setChecked(1)

            self.ui.rate_input.setText(rate)
            self.ui.event_location_input.setText(location)
            self.ui.event_location_address_input.setText(location_address)
            self.ui.event_place_combobox.setCurrentText(place)
            if lead == 'כן':
                self.ui.lead_check.setChecked(1)

        except:
            pass

        self.lead_check_btn()
        self.ui.edit_client_btn.hide()
        self.ui.update_client_btn.show()
        self.ui.update_client_btn.raise_()

    def update_customer(self):
        global current_customer
        global customer_oid

        event_date = make_date_standard(self.ui.event_date_input.toPlainText())

        current_customer.append(self.ui.event_name_input.toPlainText())
        current_customer.append(event_date)
        current_customer.append(self.ui.contact1_name_input.toPlainText())
        current_customer.append(self.ui.contact1_address_input.toPlainText())
        current_customer.append(self.ui.contact1_phone_input.toPlainText())
        current_customer.append(self.ui.contact1_email_input.toPlainText())

        if self.ui.contract_checkbox.isChecked():
            current_customer.append('קיים')
        else:
            current_customer.append('לא קיים')

        current_customer.append(self.ui.rate_input.toPlainText())
        current_customer.append(self.ui.event_location_input.toPlainText())
        current_customer.append(self.ui.event_location_address_input.toPlainText())
        current_customer.append(self.ui.event_place_combobox.currentText())

        if self.ui.lead_check.isChecked():
            current_customer.append('כן')
        else:
            current_customer.append('לא')

        current_customer.append(customer_oid)

        DataBase.update_db(self, 'customers')

        self.ui.edit_client_btn.show()
        self.ui.update_client_btn.hide()
        self.ui.edit_client_btn.raise_()
        self.load_customers_tree()
        self.ui.clear_client_fields_btn.setEnabled(True)
        self.ui.add_client_btn.setEnabled(True)
        self.ui.delete_all_clients_btn.setEnabled(True)
        self.ui.delete_selection_client_btn.setEnabled(True)

    def check_selection(self):
        btn_name = self.sender()
        row = self.ui.customers_tablewidget.currentRow()
        if 'selection' in str(btn_name) and row != -1:
            self.ui.delete_selection_client_btn.hide()
            self.ui.single_select_delete_yes_btn.show()
            self.ui.single_select_delete_no_btn.show()
            self.ui.single_select_delete_label.show()
        elif row != -1:
            self.ui.delete_all_clients_btn.hide()
            self.ui.all_select_delete_yes_btn.show()
            self.ui.all_select_delete_no_btn.show()
            self.ui.all_select_delete_label.show()

    def cancel_remove(self):
        self.ui.single_select_delete_yes_btn.hide()
        self.ui.single_select_delete_no_btn.hide()
        self.ui.single_select_delete_label.hide()
        self.ui.delete_selection_client_btn.show()
        self.ui.all_select_delete_yes_btn.hide()
        self.ui.all_select_delete_no_btn.hide()
        self.ui.all_select_delete_label.hide()
        self.ui.delete_all_clients_btn.show()

    def remove_customer(self):
        try:
            row = self.ui.customers_tablewidget.currentRow()
            column = 0
            event_name = self.ui.customers_tablewidget.item(row, column).text()
            contact_name = self.ui.customers_tablewidget.item(row, column + 2).text()

            DataBase.delete_from_db(self, 'customers', 'one', event_name, contact_name)
            self.load_customers_tree()
            personal_folders('delete', event_name)

        except:
            self.clear_all()

    def remove_all_customers(self):
        DataBase.delete_from_db(self, 'customers', 'all')
        self.load_customers_tree()
        personal_folders('delete', 'all')

    def clear_all(self):
        self.ui.event_name_input.clear()
        self.ui.event_date_input.clear()
        self.ui.contact1_name_input.clear()
        self.ui.contact1_address_input.clear()
        self.ui.contact1_phone_input.clear()
        self.ui.contact1_email_input.clear()
        self.ui.contact2_name_input.clear()
        self.ui.contact2_address_input.clear()
        self.ui.contact2_phone_input.clear()
        self.ui.contact2_email_input.clear()
        self.ui.contract_checkbox.setChecked(0)
        self.ui.rate_input.clear()
        self.ui.event_location_input.clear()
        self.ui.event_location_address_input.clear()
        self.ui.event_place_combobox.setCurrentIndex(0)
        self.ui.lead_check.setChecked(0)
        self.ui.editing_state_label.setStyleSheet(u"font: 16pt \"Segoe UI Semibold\";\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "background-color: rgb(137,236,218);\n"
                                                  "border-radius: 20px;\n"
                                                  "")
        self.ui.editing_state_label.setText('מצב הוספת אירוע')
        self.ui.event_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                               "border: 3px solid #6248F0;\n"
                                               "background-color: rgb(255,255,255);\n"
                                               "border-radius: 20px;")
        self.ui.contact1_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                  "border: 3px solid #6248F0;\n"
                                                  "background-color: rgb(255,255,255);\n"
                                                  "border-radius: 20px;")
        self.ui.event_mendatory_label.hide()
        self.ui.contact_mendatory_label.hide()

        self.ui.single_select_delete_yes_btn.hide()
        self.ui.single_select_delete_no_btn.hide()
        self.ui.single_select_delete_label.hide()
        self.ui.delete_selection_client_btn.show()
        self.ui.all_select_delete_yes_btn.hide()
        self.ui.all_select_delete_no_btn.hide()
        self.ui.all_select_delete_label.hide()
        self.ui.delete_all_clients_btn.show()

    def pop(self):
        self.ui.edit_client_btn.hide()
        self.tree_setup()
        self.ui.update_client_btn.hide()
        self.ui.contact_mendatory_label.hide()
        self.ui.event_mendatory_label.hide()
        self.ui.delete_selection_warning.hide()
        self.ui.delete_all_warning.hide()
        self.ui.excel_warning.hide()
        self.ui.clear_client_fields_btn.setEnabled(True)
        self.ui.editing_state_label.setStyleSheet(u"font: 16pt \"Segoe UI Semibold\";\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "background-color: rgb(137,236,218);\n"
                                                  "border-radius: 20px;\n"
                                                  "")
        self.ui.single_select_delete_yes_btn.hide()
        self.ui.single_select_delete_no_btn.hide()
        self.ui.single_select_delete_label.hide()
        self.ui.delete_selection_client_btn.show()
        self.ui.all_select_delete_yes_btn.hide()
        self.ui.all_select_delete_no_btn.hide()
        self.ui.all_select_delete_label.hide()
        self.ui.delete_all_clients_btn.show()
        self.show()


# SUPPLIERS WINDOW
class SuppliersWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        self.ui = Ui_SuppliersWindow()
        self.ui.setupUi(self)
        self.setWindowIcon(QtGui.QIcon('Pheonix.ico'))

        # BUTTONS
        self.ui.add_supplier_btn.clicked.connect(self.add_supplier)
        self.ui.edit_supplier_btn.clicked.connect(self.edit_supplier)
        self.ui.update_supplier_btn.clicked.connect(self.update_supplier)

        self.ui.single_select_delete_yes_btn.clicked.connect(self.remove_supplier)
        self.ui.single_select_delete_no_btn.clicked.connect(self.cancel_remove)
        self.ui.delete_selection_supplier_btn.clicked.connect(self.check_selection)
        self.ui.all_select_delete_yes_btn.clicked.connect(self.remove_all_suppliers)
        self.ui.all_select_delete_no_btn.clicked.connect(self.cancel_remove)
        self.ui.delete_all_suppliers_btn.clicked.connect(self.check_selection)

        self.ui.clear_supplier_fields_btn.clicked.connect(self.clear_all)
        self.ui.suppliers_export_excel_btn.clicked.connect(self.export_to_excel)

        # LABELS
        self.ui.delete_selection_supplier_btn.enterEvent = lambda e: self.ui.delete_selection_warning.show()
        self.ui.delete_selection_supplier_btn.leaveEvent = lambda e: self.ui.delete_selection_warning.hide()
        self.ui.delete_all_suppliers_btn.enterEvent = lambda e: self.ui.delete_all_warning.show()
        self.ui.delete_all_suppliers_btn.leaveEvent = lambda e: self.ui.delete_all_warning.hide()

    def export_to_excel(self):
        self.ui.excel_warning.hide()
        btn_clicked = self.sender()
        try:
            create_excel(btn_clicked)
        except:
            self.ui.excel_warning.show()

    def tree_setup(self):
        # TREEWIDGET VIEW
        self.ui.suppliers_tablewidget.setColumnWidth(0, 220)  # company_name
        self.ui.suppliers_tablewidget.setColumnWidth(1, 130)  # contact_name
        self.ui.suppliers_tablewidget.setColumnWidth(2, 105)  # contact_phone
        self.ui.suppliers_tablewidget.setColumnWidth(3, 190)  # contact_email
        self.ui.suppliers_tablewidget.setColumnWidth(4, 60)  # contract
        self.ui.suppliers_tablewidget.setColumnWidth(5, 320)  # extra details

        self.load_suppliers_tree()

    def load_suppliers_tree(self):
        row = 0

        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("SELECT *,oid FROM suppliers")
        suppliers = cursor.fetchall()
        cursor.close()
        db_connect.close()

        self.ui.suppliers_tablewidget.setRowCount(len(suppliers))

        for supplier in suppliers:
            company, company_address, contact_name, contact_phone, contact_email, \
            details, contract, uid = supplier  ## set up variables from tuple

            self.ui.suppliers_tablewidget.setItem(row, 0,
                                                  QtWidgets.QTableWidgetItem(company))
            self.ui.suppliers_tablewidget.setItem(row, 1,
                                                  QtWidgets.QTableWidgetItem(contact_name))
            self.ui.suppliers_tablewidget.setItem(row, 2,
                                                  QtWidgets.QTableWidgetItem(contact_phone))
            self.ui.suppliers_tablewidget.setItem(row, 3,
                                                  QtWidgets.QTableWidgetItem(contact_email))
            self.ui.suppliers_tablewidget.setItem(row, 4,
                                                  QtWidgets.QTableWidgetItem(contract))
            self.ui.suppliers_tablewidget.setItem(row, 5,
                                                  QtWidgets.QTableWidgetItem(details))
            row += 1

        self.ui.suppliers_tablewidget.resizeRowsToContents()
        if suppliers:
            self.ui.edit_supplier_btn.show()
        else:
            self.ui.edit_supplier_btn.hide()
        self.clear_all()

    def add_supplier(self):

        if self.ui.company_name_input.toPlainText() != '' and self.ui.contact1_name_input.toPlainText() != '':
            # CheckBox for contract
            if self.ui.arrangment_checkbox.isChecked():
                contract = 'קיים'
            else:
                contract = 'לא קיים'

            supplier_1 = {'company': self.ui.company_name_input.toPlainText(),
                          'company_address': self.ui.company_address_input.toPlainText(),
                          'contact_name': self.ui.contact1_name_input.toPlainText(),
                          'contact_phone': self.ui.contact1_phone_input.toPlainText(),
                          'contact_email': self.ui.contact1_email_input.toPlainText(),
                          'contract': contract,
                          'details': self.ui.details_input.toPlainText(),
                          }

            DataBase.insert_into_db(self, 'suppliers', supplier_1)

            if self.ui.contact2_name_input.toPlainText() != '':
                supplier_2 = {'company': self.ui.company_name_input.toPlainText(),
                              'company_address': self.ui.company_address_input.toPlainText(),
                              'contact_name': self.ui.contact2_name_input.toPlainText(),
                              'contact_phone': self.ui.contact2_phone_input.toPlainText(),
                              'contact_email': self.ui.contact2_email_input.toPlainText(),
                              'contract': contract,
                              'details': self.ui.details_input.toPlainText(),
                              }

                DataBase.insert_into_db(self, 'suppliers', supplier_2)

            self.load_suppliers_tree()

        elif self.ui.company_name_input.toPlainText() == '' and self.ui.contact1_name_input.toPlainText() != '':
            self.ui.company_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                     "border: 3px solid #ef4040;\n"
                                                     "background-color: rgb(255,255,255);\n"
                                                     "border-radius: 20px;")
            self.ui.company_mendatory_label.show()
            self.ui.company_mendatory_label.raise_()

        elif self.ui.contact1_name_input.toPlainText() == '' and self.ui.company_name_input.toPlainText() != '':
            self.ui.contact1_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                      "border: 3px solid #ef4040;\n"
                                                      "background-color: rgb(255,255,255);\n"
                                                      "border-radius: 20px;")
            self.ui.contact_mendatory_label.show()
            self.ui.contact_mendatory_label.raise_()

        else:
            self.ui.company_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                     "border: 3px solid #ef4040;\n"
                                                     "background-color: rgb(255,255,255);\n"
                                                     "border-radius: 20px;")

            self.ui.contact1_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                      "border: 3px solid #ef4040;\n"
                                                      "background-color: rgb(255,255,255);\n"
                                                      "border-radius: 20px;")
            self.ui.contact_mendatory_label.show()
            self.ui.company_mendatory_label.show()
            self.ui.company_mendatory_label.raise_()
            self.ui.contact_mendatory_label.raise_()

    def edit_supplier(self):
        global supplier_oid
        self.ui.clear_supplier_fields_btn.setEnabled(False)
        self.ui.add_supplier_btn.setEnabled(False)
        self.ui.delete_all_suppliers_btn.setEnabled(False)
        self.ui.delete_selection_supplier_btn.setEnabled(False)
        self.ui.contact_mendatory_label.hide()
        self.ui.company_mendatory_label.hide()

        try:
            row = self.ui.suppliers_tablewidget.currentRow()
            column = 0
            company = self.ui.suppliers_tablewidget.item(row, column).text()
            contact_name = self.ui.suppliers_tablewidget.item(row, column + 1).text()

            record = DataBase.fetch_from_db(self, 'suppliers', 'one', company, contact_name)

            company, company_address, contact_name, contact_phone, contact_email, details, \
            contract, supplier_oid = record[0]  ## set up variables from tuple

            self.ui.company_name_input.setText(company)
            self.ui.company_address_input.setText(company_address)
            self.ui.contact1_name_input.setText(contact_name)
            self.ui.contact1_phone_input.setText(contact_phone)
            self.ui.contact1_email_input.setText(contact_email)
            self.ui.details_input.setText(details)

            if contract == 'קיים':
                self.ui.arrangment_checkbox.setChecked(1)

            self.ui.edit_supplier_btn.hide()
            self.ui.update_supplier_btn.show()
            self.ui.update_supplier_btn.raise_()

        except:
            pass

    def update_supplier(self):
        global current_supplier
        global supplier_oid

        current_supplier.append(self.ui.company_name_input.toPlainText())
        current_supplier.append(self.ui.company_address_input.toPlainText())
        current_supplier.append(self.ui.contact1_name_input.toPlainText())
        current_supplier.append(self.ui.contact1_phone_input.toPlainText())
        current_supplier.append(self.ui.contact1_email_input.toPlainText())
        current_supplier.append(self.ui.details_input.toPlainText())

        if self.ui.arrangment_checkbox.isChecked():
            current_supplier.append('קיים')
        else:
            current_supplier.append('לא קיים')

        current_supplier.append(supplier_oid)

        DataBase.update_db(self, 'suppliers')

        self.ui.edit_supplier_btn.show()
        self.ui.update_supplier_btn.hide()
        self.ui.edit_supplier_btn.raise_()
        self.load_suppliers_tree()
        self.ui.clear_supplier_fields_btn.setEnabled(True)
        self.ui.add_supplier_btn.setEnabled(True)
        self.ui.delete_all_suppliers_btn.setEnabled(True)
        self.ui.delete_selection_supplier_btn.setEnabled(True)

    def check_selection(self):
        btn_name = self.sender()
        row = self.ui.suppliers_tablewidget.currentRow()
        if 'selection' in str(btn_name) and row != -1:
            self.ui.delete_selection_supplier_btn.hide()
            self.ui.single_select_delete_yes_btn.show()
            self.ui.single_select_delete_no_btn.show()
            self.ui.single_select_delete_label.show()
        elif row != -1:
            self.ui.delete_all_suppliers_btn.hide()
            self.ui.all_select_delete_yes_btn.show()
            self.ui.all_select_delete_no_btn.show()
            self.ui.all_select_delete_label.show()

    def cancel_remove(self):
        self.ui.single_select_delete_yes_btn.hide()
        self.ui.single_select_delete_no_btn.hide()
        self.ui.single_select_delete_label.hide()
        self.ui.delete_selection_supplier_btn.show()
        self.ui.all_select_delete_yes_btn.hide()
        self.ui.all_select_delete_no_btn.hide()
        self.ui.all_select_delete_label.hide()
        self.ui.delete_all_suppliers_btn.show()

    def remove_supplier(self):
        try:
            row = self.ui.suppliers_tablewidget.currentRow()
            column = 0
            company = self.ui.suppliers_tablewidget.item(row, column).text()
            contact1_name = self.ui.suppliers_tablewidget.item(row, column + 1).text()

            DataBase.delete_from_db(self, 'suppliers', 'one', company, contact1_name)  # delete contact 1
            self.load_suppliers_tree()

        except:
            self.clear_all()

    def remove_all_suppliers(self):
        DataBase.delete_from_db(self, 'suppliers', 'all')
        self.load_suppliers_tree()

    def clear_all(self):
        self.ui.company_name_input.clear()
        self.ui.company_address_input.clear()
        self.ui.contact1_name_input.clear()
        self.ui.contact1_phone_input.clear()
        self.ui.contact1_email_input.clear()
        self.ui.contact2_name_input.clear()
        self.ui.contact2_phone_input.clear()
        self.ui.contact2_email_input.clear()
        self.ui.details_input.clear()
        self.ui.arrangment_checkbox.setChecked(0)
        self.ui.company_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                 "border: 3px solid #6248F0;\n"
                                                 "background-color: rgb(255,255,255);\n"
                                                 "border-radius: 20px;")
        self.ui.contact1_name_input.setStyleSheet(u"font: 11pt \"Segoe UI Semilight\";\n"
                                                  "border: 3px solid #6248F0;\n"
                                                  "background-color: rgb(255,255,255);\n"
                                                  "border-radius: 20px;")
        self.ui.company_mendatory_label.hide()
        self.ui.contact_mendatory_label.hide()

        self.ui.single_select_delete_yes_btn.hide()
        self.ui.single_select_delete_no_btn.hide()
        self.ui.single_select_delete_label.hide()
        self.ui.delete_selection_supplier_btn.show()
        self.ui.all_select_delete_yes_btn.hide()
        self.ui.all_select_delete_no_btn.hide()
        self.ui.all_select_delete_label.hide()
        self.ui.delete_all_suppliers_btn.show()

    def pop(self):
        self.ui.edit_supplier_btn.hide()
        self.tree_setup()
        self.ui.update_supplier_btn.hide()
        self.ui.company_mendatory_label.hide()
        self.ui.contact_mendatory_label.hide()
        self.ui.delete_selection_warning.hide()
        self.ui.delete_all_warning.hide()
        self.ui.excel_warning.hide()
        self.ui.single_select_delete_yes_btn.hide()
        self.ui.single_select_delete_no_btn.hide()
        self.ui.single_select_delete_label.hide()
        self.ui.delete_selection_supplier_btn.show()
        self.ui.all_select_delete_yes_btn.hide()
        self.ui.all_select_delete_no_btn.hide()
        self.ui.all_select_delete_label.hide()
        self.ui.delete_all_suppliers_btn.show()
        self.ui.clear_supplier_fields_btn.setEnabled(True)
        self.show()


class LeadWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_LeadWindow()
        self.ui.setupUi(self)
        self.setWindowIcon(QtGui.QIcon('Pheonix.ico'))

        # VARIABLES
        self.folder_path = 'לידים/'
        if not os.path.isdir(self.folder_path):
            os.mkdir(self.folder_path)

        # BUTTONS
        self.ui.lead_folder_btn.clicked.connect(self.open_folder)

    def open_folder(self):
        self.lead_name = self.ui.upper_frame_header_btn.text()
        webbrowser.open(os.path.realpath(self.folder_path + self.lead_name))

    def pop_window(self, btn_name):
        self.ui.upper_frame_header_btn.setText(btn_name)
        self.event_name = self.ui.upper_frame_header_btn.text()

        if not os.path.isdir(self.folder_path + btn_name):
            os.mkdir(self.folder_path + btn_name)
            personal_folders('create', btn_name)

        # self.show()

    def add_files(self):
        pass

    def pop_alerts(self, btn_name):
        startIndex = str(btn_name).find('\"') + 1
        endIndex = str(btn_name).find('\"', startIndex + 1)
        btn_name = str(btn_name)[startIndex:endIndex]

        # self.show()


class EventWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_EventWindow()
        self.ui.setupUi(self)
        self.load_dragndrop()
        self.setWindowIcon(QtGui.QIcon('Pheonix.ico'))

        # BUTTONS
        self.ui.event_collapse_all_btn.clicked.connect(self.collapse_all)
        self.ui.event_clear_all_btn.clicked.connect(self.clear_all)
        self.ui.browser_btn.clicked.connect(self.browse_files)

        self.ui.client_details_btn.clicked.connect(self.show_client_details)
        self.ui.event_details_btn.clicked.connect(self.show_event_details)
        self.ui.update_budget_btn.clicked.connect(self.update_budget)
        self.ui.adding_btn.clicked.connect(self.show_add_buttons)

        self.ui.event_folder_btn.clicked.connect(self.open_folder)
        self.ui.event_folder_btn.enterEvent = lambda e: self.ui.event_folder_label.show()
        self.ui.event_folder_btn.leaveEvent = lambda e: self.ui.event_folder_label.hide()

        self.ui.event_excel_btn.clicked.connect(self.export_to_excel)
        self.ui.event_excel_btn.enterEvent = lambda e: self.ui.event_excel_label.show()
        self.ui.event_excel_btn.leaveEvent = lambda e: self.ui.event_excel_label.hide()

        self.ui.event_delete_btn.clicked.connect(self.delete_event)
        self.ui.event_delete_btn.enterEvent = lambda e: self.ui.event_delete_label.show()
        self.ui.event_delete_btn.leaveEvent = lambda e: self.ui.event_delete_label.hide()

        # STACKED WIDGETS
        self.ui.events_stackedWidget.setCurrentWidget(self.ui.files_page)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ PAYMENTS PAGE ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.ui.payments_page_btn.clicked.connect(self.switchto_payments_page)
        self.ui.delete_all_payments_btn.enterEvent = lambda e: self.ui.payment_delete_all_warning.show()
        self.ui.delete_all_payments_btn.leaveEvent = lambda e: self.ui.payment_delete_all_warning.hide()
        self.ui.delete_selection_payment_btn.enterEvent = lambda e: self.ui.payment_delete_selection_warning.show()
        self.ui.delete_selection_payment_btn.leaveEvent = lambda e: self.ui.payment_delete_selection_warning.hide()

        self.ui.add_payment_btn.clicked.connect(self.add_payment)
        self.ui.edit_payment_btn.clicked.connect(self.edit_payment)
        self.ui.update_payment_btn.clicked.connect(self.update_payment)
        self.ui.delete_selection_payment_btn.clicked.connect(self.check_payment_selection)
        self.ui.delete_all_payments_btn.clicked.connect(self.check_payment_selection)

        self.ui.payments_all_select_delete_yes_btn.clicked.connect(self.remove_all_payments)
        self.ui.payments_all_select_delete_no_btn.clicked.connect(self.cancel_payment_remove)
        self.ui.payment_single_select_delete_yes_btn.clicked.connect(self.remove_payment)
        self.ui.payment_single_select_delete_no_btn.clicked.connect(self.cancel_payment_remove)

        self.ui.payments_export_excel_btn.clicked.connect(self.export_to_excel)

        # SUPPLIER COMBO-BOX
        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("SELECT DISTINCT company FROM suppliers")
        suppliers_names = cursor.fetchall()
        cursor.close()
        db_connect.close()
        for supplier in suppliers_names:
            self.ui.suppliers_combobox.addItems(supplier)

        self.ui.suppliers_combobox.currentTextChanged.connect(self.show_guest_info)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ MISSIONS PAGE ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.ui.missions_page_btn.clicked.connect(self.switchto_missions_page)

        self.ui.delete_all_missions_btn.enterEvent = lambda e: self.ui.missions_delete_all_warning.show()
        self.ui.delete_all_missions_btn.leaveEvent = lambda e: self.ui.missions_delete_all_warning.hide()
        self.ui.delete_selection_mission_btn.enterEvent = lambda e: self.ui.mission_delete_selection_warning.show()
        self.ui.delete_selection_mission_btn.leaveEvent = lambda e: self.ui.mission_delete_selection_warning.hide()

        self.ui.add_mission_btn.clicked.connect(self.add_mission)
        self.ui.edit_mission_btn.clicked.connect(self.edit_mission)
        self.ui.update_mission_btn.clicked.connect(self.update_mission)
        self.ui.delete_selection_mission_btn.clicked.connect(self.check_mission_selection)
        self.ui.delete_all_missions_btn.clicked.connect(self.check_mission_selection)

        self.ui.missions_all_select_delete_yes_btn.clicked.connect(self.remove_all_missions)
        self.ui.missions_all_select_delete_no_btn.clicked.connect(self.cancel_mission_remove)
        self.ui.mission_single_select_delete_yes_btn.clicked.connect(self.remove_mission)
        self.ui.mission_single_select_delete_no_btn.clicked.connect(self.cancel_mission_remove)

        self.ui.missions_export_excel_btn.clicked.connect(self.export_to_excel)

        # STATUS COMBO-BOX
        self.ui.status_combobox.currentTextChanged.connect(self.status_color_change)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ FILES PAGE ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.ui.files_page_btn.clicked.connect(self.switchto_files_page)

        # VARIABLES
        self.folder_path = 'הפקות/'
        if not os.path.isdir(self.folder_path):
            os.mkdir(self.folder_path)

    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ PAYMENTS TREE ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def payments_tree_setup(self):
        self.ui.payments_tablewidget.setColumnWidth(0, 250)  # service
        self.ui.payments_tablewidget.setColumnWidth(1, 175)  # supplier
        self.ui.payments_tablewidget.setColumnWidth(2, 60)  # price
        self.ui.payments_tablewidget.setColumnWidth(3, 65)  # advance
        self.ui.payments_tablewidget.setColumnWidth(4, 70)  # left to pay
        self.ui.payments_tablewidget.setColumnWidth(5, 45)  # tip
        self.ui.payments_tablewidget.setColumnWidth(6, 310)  # details
        self.load_payments_tree()

    def load_payments_tree(self):
        row = 0

        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("SELECT *,oid FROM payments WHERE event_name=:event_name",
                       {'event_name': self.event_name
                        })
        payments = cursor.fetchall()
        cursor.close()
        db_connect.close()

        self.ui.payments_tablewidget.setRowCount(len(payments))

        for payment in payments:
            event_name, category, service, supplier, price, advance, \
            advance_status, tip, details, payment_oid = payment  ## set up variables from tuple

            if '~~~' in details:
                if str(advance) != '' and advance_status == 'כן':
                    to_pay = str(int(price) - int(advance))
                else:
                    to_pay = str(price)
            else:
                if str(price) != '' and str(advance) != '' and advance_status == 'כן':
                    to_pay = str(int(price) - int(advance))
                else:
                    to_pay = str(price)

            price = make_amount_readable(price)
            advance = make_amount_readable(advance)
            to_pay = make_amount_readable(to_pay)
            tip = make_amount_readable(tip)

            self.ui.payments_tablewidget.setItem(row, 0,
                                                 QtWidgets.QTableWidgetItem(service))
            self.ui.payments_tablewidget.setItem(row, 1,
                                                 QtWidgets.QTableWidgetItem(supplier))
            self.ui.payments_tablewidget.setItem(row, 2,
                                                 QtWidgets.QTableWidgetItem(price))
            self.ui.payments_tablewidget.setItem(row, 3,
                                                 QtWidgets.QTableWidgetItem(advance))
            self.ui.payments_tablewidget.setItem(row, 4,
                                                 QtWidgets.QTableWidgetItem(to_pay))
            self.ui.payments_tablewidget.setItem(row, 5,
                                                 QtWidgets.QTableWidgetItem(tip))
            self.ui.payments_tablewidget.setItem(row, 6,
                                                 QtWidgets.QTableWidgetItem(details))
            row += 1

        self.ui.payments_tablewidget.resizeRowsToContents()
        if self.ui.payments_tablewidget.rowCount() > 0:
            self.ui.edit_payment_btn.show()
        self.clear_text()

    def add_payment(self):
        category = self.ui.category_combobox.currentText()
        if self.ui.suppliers_combobox.currentText() != 'בחר ספק' and self.ui.supplier_service_text.toPlainText() != '':
            # CheckBox for advance
            if self.ui.advance_checkbox.isChecked():
                advance_status = 'כן'
            else:
                advance_status = 'לא'

            price = self.ui.supplier_price_text.toPlainText().replace(',', ''). \
                replace(' ', '').replace('-', '').replace('/', '')

            details = self.ui.supplier_details_text.toPlainText()

            if not self.ui.guest_num_input.isHidden():
                if self.ui.supplier_details_text.toPlainText() != '':
                    guests = self.ui.guest_num_input.toPlainText()
                    details_input = self.ui.supplier_details_text.toPlainText()
                    details = '~~!~~ ' + f'לפי {price} למנה' + \
                              f', {guests} מוזמנים' + ' ~~!~~' + \
                              f'\n{details_input}'
                else:
                    details = '~~!~~ ' + f'לפי {price} למנה' + \
                              f', {self.ui.guest_num_input.toPlainText()} מוזמנים' + ' ~~!~~'

                price = str(int(self.ui.supplier_price_text.toPlainText().replace(',', '').replace(' ', '')
                                .replace('-', '').replace('/', '')) * int(self.ui.guest_num_input.toPlainText()))

            payment = {'event_name': self.event_name,
                       'category': category,
                       'service': self.ui.supplier_service_text.toPlainText(),
                       'supplier': self.ui.suppliers_combobox.currentText(),
                       'price': price,
                       'advance': self.ui.supplier_advance_text.toPlainText(),
                       'advance_status': advance_status,
                       'tip': self.ui.supplier_tip_text.toPlainText(),
                       'details': details
                       }

            # add to Categories DB if category does not exist in it
            exist_flag = False
            categories = DataBase.fetch_from_db(self, 'categories')
            for item in categories:
                if category == item[0]:
                    exist_flag = True
            if not exist_flag:
                DataBase.insert_into_db(self, 'categories', category)

            DataBase.insert_into_db(self, 'payments', payment)
            self.subtitle_budget_status()
            self.load_payments_tree()
            self.clear_text()

        elif self.ui.suppliers_combobox.currentText() != 'בחר ספק' and self.ui.supplier_service_text.toPlainText() == '':
            self.ui.supplier_service_text.setStyleSheet(u"font: 12pt \"Segoe UI Semilight\";\n"
                                                        "border: 3px solid #ef4040;\n"
                                                        "background-color: rgb(255,255,255);\n"
                                                        "border-radius: 20px;")
            self.ui.suppliers_combobox.setStyleSheet(u"font: 12pt \"Segoe UI Semilight\";\n"
                                                     "border: 3px solid #6248F0;\n"
                                                     "background-color: rgb(255,255,255);\n"
                                                     "border-radius: 20px;")
            self.ui.supplier_mendatory_label.hide()
            self.ui.service_mendatory_label.show()
            self.ui.service_mendatory_label.raise_()

        elif self.ui.suppliers_combobox.currentText() == 'בחר ספק' and self.ui.supplier_service_text.toPlainText() != '':
            self.ui.supplier_service_text.setStyleSheet(u"font: 12pt \"Segoe UI Semilight\";\n"
                                                        "border: 3px solid #6248F0;\n"
                                                        "background-color: rgb(255,255,255);\n"
                                                        "border-radius: 20px;")
            self.ui.suppliers_combobox.setStyleSheet(u"font: 12pt \"Segoe UI Semilight\";\n"
                                                     "border: 3px solid #ef4040;\n"
                                                     "background-color: rgb(255,255,255);\n"
                                                     "border-radius: 20px;")
            self.ui.service_mendatory_label.hide()
            self.ui.supplier_mendatory_label.show()
            self.ui.supplier_mendatory_label.raise_()

        else:
            self.ui.suppliers_combobox.setStyleSheet(u"font: 12pt \"Segoe UI Semilight\";\n"
                                                     "border: 3px solid #ef4040;\n"
                                                     "background-color: rgb(255,255,255);\n"
                                                     "border-radius: 20px;")
            self.ui.supplier_mendatory_label.show()
            self.ui.supplier_mendatory_label.raise_()

            self.ui.supplier_service_text.setStyleSheet(u"font: 12pt \"Segoe UI Semilight\";\n"
                                                        "border: 3px solid #ef4040;\n"
                                                        "background-color: rgb(255,255,255);\n"
                                                        "border-radius: 20px;")
            self.ui.service_mendatory_label.show()
            self.ui.service_mendatory_label.raise_()

    def check_payment_selection(self):
        btn_name = self.sender()
        row = self.ui.payments_tablewidget.currentRow()
        if 'selection' in str(btn_name) and row != -1:
            self.ui.delete_selection_payment_btn.hide()
            self.ui.payment_single_select_delete_yes_btn.show()
            self.ui.payment_single_select_delete_no_btn.show()
            self.ui.payment_single_select_delete_label.show()
        elif row != -1:
            self.ui.delete_all_payments_btn.hide()
            self.ui.payments_all_select_delete_yes_btn.show()
            self.ui.payments_all_select_delete_no_btn.show()
            self.ui.payments_all_select_delete_label.show()

    def cancel_payment_remove(self):
        self.ui.delete_selection_payment_btn.show()
        self.ui.payment_single_select_delete_yes_btn.hide()
        self.ui.payment_single_select_delete_no_btn.hide()
        self.ui.payment_single_select_delete_label.hide()
        self.ui.delete_all_payments_btn.show()
        self.ui.payments_all_select_delete_yes_btn.hide()
        self.ui.payments_all_select_delete_no_btn.hide()
        self.ui.payments_all_select_delete_label.hide()

    def edit_payment(self):
        global payment_oid
        self.ui.clear_payment_text_btn.setEnabled(False)
        self.ui.add_payment_btn.setEnabled(False)
        self.ui.delete_all_payments_btn.setEnabled(False)
        self.ui.delete_selection_payment_btn.setEnabled(False)
        self.ui.supplier_mendatory_label.hide()
        self.ui.service_mendatory_label.hide()

        try:
            row = self.ui.payments_tablewidget.currentRow()
            column = 0
            event_name = self.event_name
            service = self.ui.payments_tablewidget.item(row, column).text()
            supplier = self.ui.payments_tablewidget.item(row, column + 1).text()

            record = DataBase.fetch_from_db(self, 'payments', 'one', event_name, supplier, service)

            event_name, category, service, supplier, price, advance, advance_status, \
            tip, details, payment_oid = record[0]  ## set up variables from tuple

            if '~~' in details:
                self.ui.guest_num_input.setText(details.split(',')[1].split(' מוזמנים')[0].split(' ')[1])
                price = details.split('לפי ')[1].split(' למנה')[0]
                self.ui.guest_num_input.show()

            self.ui.category_combobox.setCurrentText(category)
            self.ui.suppliers_combobox.setCurrentText(supplier)
            self.ui.supplier_service_text.setText(service)
            self.ui.supplier_price_text.setText(price)
            self.ui.supplier_advance_text.setText(advance)

            if advance_status == 'כן':
                self.ui.advance_checkbox.setChecked(1)

            self.ui.supplier_tip_text.setText(tip)
            self.ui.supplier_details_text.setText(details)

            self.ui.edit_payment_btn.hide()
            self.ui.update_payment_btn.show()
            self.ui.update_payment_btn.raise_()
        except Exception as e:
            pass

    def update_payment(self):
        global current_payment
        global payment_oid
        category = self.ui.category_combobox.currentText()

        current_payment.append(self.event_name)
        current_payment.append(self.ui.category_combobox.currentText())
        current_payment.append(self.ui.supplier_service_text.toPlainText())
        current_payment.append(self.ui.suppliers_combobox.currentText())

        price = self.ui.supplier_price_text.toPlainText().replace(',', '').replace(' ', '') \
            .replace('-', '').replace('/', '')

        details = self.ui.supplier_details_text.toPlainText()

        if not self.ui.guest_num_input.isHidden():
            if self.ui.supplier_details_text.toPlainText() != '' \
                    and '~~' not in self.ui.supplier_details_text.toPlainText():
                details = '~~!~~ ' + f'לפי {price} למנה' + \
                          f', {self.ui.guest_num_input.toPlainText()} מוזמנים' + \
                          ' ~~!~~' + f'\n{self.ui.supplier_details_text.toPlainText()}'

            elif '~~' in self.ui.supplier_details_text.toPlainText():
                info = self.ui.supplier_details_text.toPlainText()
                extra_info = "".join((info[info.rfind('~') + 1:], ""))
                details = "".join((info[:info.rfind('~') + 1], extra_info))

            price = str(int(self.ui.supplier_price_text.toPlainText().replace(',', '').replace(' ', '') \
                            .replace('-', '').replace('/', '')) * int(self.ui.guest_num_input.toPlainText()))

        current_payment.append(price)
        current_payment.append(self.ui.supplier_advance_text.toPlainText())

        if self.ui.advance_checkbox.isChecked():
            current_payment.append('כן')
        else:
            current_payment.append('לא')

        current_payment.append(self.ui.supplier_tip_text.toPlainText())

        current_payment.append(details)
        current_payment.append(payment_oid)

        DataBase.update_db(self, 'payments')

        # add to Categories DB if category does not exist in it
        exist_flag = False
        categories = DataBase.fetch_from_db(self, 'categories')
        for item in categories:
            if category == item[0]:
                exist_flag = True
        if not exist_flag:
            DataBase.insert_into_db(self, 'categories', category)

        self.ui.edit_payment_btn.show()
        self.ui.update_payment_btn.hide()
        self.ui.edit_payment_btn.raise_()
        self.load_payments_tree()
        self.ui.clear_payment_text_btn.setEnabled(True)
        self.ui.add_payment_btn.setEnabled(True)
        self.ui.delete_all_payments_btn.setEnabled(True)
        self.ui.delete_selection_payment_btn.setEnabled(True)

    def remove_payment(self):
        try:
            row = self.ui.payments_tablewidget.currentRow()
            column = 0
            event_name = self.event_name
            service = self.ui.payments_tablewidget.item(row, column).text()
            supplier = self.ui.payments_tablewidget.item(row, column + 1).text()

            DataBase.delete_from_db(self, 'payments', 'one', event_name, service, supplier)
            self.load_payments_tree()
            self.subtitle_budget_status()

        except:
            self.clear_text()

        self.ui.delete_selection_payment_btn.show()
        self.ui.payment_single_select_delete_yes_btn.hide()
        self.ui.payment_single_select_delete_no_btn.hide()
        self.ui.payment_single_select_delete_label.hide()

    def remove_all_payments(self):
        DataBase.delete_from_db(self, 'payments', 'all', self.event_name)
        self.subtitle_budget_status()

        self.ui.delete_all_payments_btn.show()
        self.ui.payments_all_select_delete_yes_btn.hide()
        self.ui.payments_all_select_delete_no_btn.hide()
        self.ui.payments_all_select_delete_label.hide()
        self.load_payments_tree()

    def show_guest_info(self, value):
        if value == 'מחיר מנה - אולם':
            self.ui.supplier_price_text.setPlaceholderText('מחיר למנה')
            self.ui.guest_num_input.show()
        else:
            self.ui.supplier_price_text.setPlaceholderText(
                u"\u05de\u05d7\u05d9\u05e8 \u05db\u05d5\u05dc\u05dc \u05de\u05e2\"\u05de")
            self.ui.guest_num_input.hide()

    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ MISSIONS TREE ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    def missions_tree_setup(self):
        self.ui.missions_tablewidget.setColumnWidth(0, 385)  # mission
        self.ui.missions_tablewidget.setColumnWidth(1, 120)  # due date
        self.ui.missions_tablewidget.setColumnWidth(2, 80)  # status
        self.ui.missions_tablewidget.setColumnWidth(3, 400)  # notes
        self.load_missions_tree()

    def load_missions_tree(self):
        row = 0
        status_color = {'בתהליך': QtGui.QColor(255, 223, 128),
                        'בוצע': QtGui.QColor(165, 212, 106),
                        'לא רלוונטי': QtGui.QColor(227, 140, 140)}

        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("SELECT *,oid FROM missions WHERE event_name=:event_name ORDER BY mission_due",
                       {'event_name': self.event_name
                        })
        missions = cursor.fetchall()
        cursor.close()
        db_connect.close()

        self.ui.missions_tablewidget.setRowCount(len(missions))

        for item in missions:
            event_name, mission, mission_due, status, notes, mission_oid = item  ## set up variables from tuple

            self.ui.missions_tablewidget.setItem(row, 0,
                                                 QtWidgets.QTableWidgetItem(mission))
            self.ui.missions_tablewidget.setItem(row, 1,
                                                 QtWidgets.QTableWidgetItem(mission_due))
            self.ui.missions_tablewidget.setItem(row, 2,
                                                 QtWidgets.QTableWidgetItem(status))
            self.ui.missions_tablewidget.setItem(row, 3,
                                                 QtWidgets.QTableWidgetItem(notes))

            self.ui.missions_tablewidget.item(row, 2).setBackground(status_color[status])
            row += 1

        self.ui.missions_tablewidget.resizeRowsToContents()
        if self.ui.missions_tablewidget.rowCount() > 0:
            self.ui.edit_mission_btn.show()
        self.clear_text()

    def add_mission(self):
        if self.ui.mission_name_text.toPlainText() != '':

            new_date = make_date_standard(self.ui.mission_due_text.toPlainText())
            mission = {'event_name': self.event_name,
                       'mission': self.ui.mission_name_text.toPlainText(),
                       'mission_due': new_date,
                       'status': self.ui.status_combobox.currentText(),
                       'notes': self.ui.mission_notes_text.toPlainText()
                       }

            DataBase.insert_into_db(self, 'missions', mission)
            self.load_missions_tree()
            self.clear_text()

        else:
            self.ui.mission_name_text.setStyleSheet(u"font: 12pt \"Segoe UI Semilight\";\n"
                                                    "border: 3px solid #ef4040;\n"
                                                    "background-color: rgb(255,255,255);\n"
                                                    "border-radius: 20px;")
            self.ui.mission_mendatory_label.show()
            self.ui.service_mendatory_label.raise_()


    def check_mission_selection(self):
        btn_name = self.sender()
        row = self.ui.missions_tablewidget.currentRow()
        if 'selection' in str(btn_name) and row != -1:
            self.ui.delete_selection_mission_btn.hide()
            self.ui.mission_single_select_delete_yes_btn.show()
            self.ui.mission_single_select_delete_no_btn.show()
            self.ui.mission_single_select_delete_label.show()
        elif row != -1:
            self.ui.delete_all_missions_btn.hide()
            self.ui.missions_all_select_delete_yes_btn.show()
            self.ui.missions_all_select_delete_no_btn.show()
            self.ui.missions_all_select_delete_label.show()

    def cancel_mission_remove(self):
        self.ui.delete_selection_mission_btn.show()
        self.ui.mission_single_select_delete_yes_btn.hide()
        self.ui.mission_single_select_delete_no_btn.hide()
        self.ui.mission_single_select_delete_label.hide()
        self.ui.delete_all_missions_btn.show()
        self.ui.missions_all_select_delete_yes_btn.hide()
        self.ui.missions_all_select_delete_no_btn.hide()
        self.ui.missions_all_select_delete_label.hide()

    def edit_mission(self):
        global mission_oid

        self.ui.clear_mission_input_btn.setEnabled(False)
        self.ui.add_mission_btn.setEnabled(False)
        self.ui.delete_all_missions_btn.setEnabled(False)
        self.ui.delete_selection_mission_btn.setEnabled(False)
        self.ui.mission_mendatory_label.hide()

        try:
            row = self.ui.missions_tablewidget.currentRow()
            column = 0
            event_name = self.event_name
            mission_data = self.ui.missions_tablewidget.item(row, column).text()

            record = DataBase.fetch_from_db(self, 'missions', 'one', event_name, mission_data)
            event_name, mission, mission_due, status, notes, mission_oid = record[0]  ## set up variables from tuple

            self.ui.mission_name_text.setText(mission)
            self.ui.mission_due_text.setText(mission_due)
            self.ui.status_combobox.setCurrentText(status)
            self.ui.mission_notes_text.setText(notes)
        except:
            pass

        self.ui.edit_mission_btn.hide()
        self.ui.update_mission_btn.show()
        self.ui.update_mission_btn.raise_()

    def update_mission(self):
        global current_mission
        global mission_oid

        new_date = make_date_standard(self.ui.mission_due_text.toPlainText())
        current_mission.append(self.event_name)
        current_mission.append(self.ui.mission_name_text.toPlainText())
        current_mission.append(new_date)
        current_mission.append(self.ui.status_combobox.currentText())
        current_mission.append(self.ui.mission_notes_text.toPlainText())
        current_mission.append(mission_oid)

        DataBase.update_db(self, 'missions')

        self.ui.edit_mission_btn.show()
        self.ui.update_mission_btn.hide()
        self.ui.edit_mission_btn.raise_()
        self.load_missions_tree()
        self.ui.clear_mission_input_btn.setEnabled(True)
        self.ui.add_mission_btn.setEnabled(True)
        self.ui.delete_all_missions_btn.setEnabled(True)
        self.ui.delete_selection_mission_btn.setEnabled(True)

    def remove_mission(self):
        try:
            row = self.ui.missions_tablewidget.currentRow()
            column = 0
            event_name = self.event_name
            mission_data = self.ui.missions_tablewidget.item(row, column).text()
            DataBase.delete_from_db(self, 'missions', 'one', event_name, mission_data)

            self.load_missions_tree()

        except:
            self.clear_text()

        self.ui.delete_selection_mission_btn.show()
        self.ui.mission_single_select_delete_yes_btn.hide()
        self.ui.mission_single_select_delete_no_btn.hide()
        self.ui.mission_single_select_delete_label.hide()

    def remove_all_missions(self):
        DataBase.delete_from_db(self, 'missions', 'all', self.event_name)
        self.ui.delete_all_missions_btn.show()
        self.ui.missions_all_select_delete_yes_btn.hide()
        self.ui.missions_all_select_delete_no_btn.hide()
        self.ui.missions_all_select_delete_label.hide()
        self.load_missions_tree()

    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ FILES and DRAGnDROP FUNCTIONS ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    def load_dragndrop(self):
        self.dragndrop_list = ListBoxWidget(self)
        QListWidgetItem(self.dragndrop_list)
        self.dragndrop_list.setGeometry(QRect(450, 460, 500, 390))
        self.dragndrop_list.setStyleSheet(u"font: 14pt \"Segoe UI Semilight\";\n"
                                          "border: 1px solid #090206;\n"
                                          "background-color: rgb(255,255,255);\n"
                                          "border-radius: 20px;")
        self.dragndrop_list.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.dragndrop_list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.dragndrop_list.setDragDropOverwriteMode(True)
        self.dragndrop_list.setDragDropMode(QAbstractItemView.DragDrop)
        self.dragndrop_list.setAlternatingRowColors(True)
        self.dragndrop_list.setTextElideMode(Qt.ElideMiddle)
        self.dragndrop_list.setItemAlignment(Qt.AlignCenter | Qt.AlignHCenter |
                                             Qt.AlignRight | Qt.AlignTrailing |
                                             Qt.AlignVCenter)
        self.dragndrop_list.raise_()

    def open_folder(self):
        self.event_name = self.ui.headline_label.text()
        webbrowser.open(os.path.realpath(self.folder_path + self.event_name))

    def browse_files(self):
        files_selected = QFileDialog.getOpenFileNames(self, 'בחר קובץ', os.getcwd())
        for file in files_selected[0]:
            self.add_files(file)

    def add_files(self, file):
        file_name = file.split("/")[-1]
        file_type = file.split("/")[-1].split(".")[-1]
        if file_type in Image_Types:
            shutil.copyfile(file, self.folder_path + "/" + self.event_name + '/' + 'תמונות וסרטונים' + '/' + file_name)
            self.dragndrop_list.addItem(file_name.split(".")[0] + ' - נוסף בהצלחה')

        elif file_type in Document_Types:
            shutil.copyfile(file, self.folder_path + "/" + self.event_name + '/' + 'מסמכים' + '/' + file_name)
            self.dragndrop_list.addItem(file_name.split(".")[0] + ' - נוסף בהצלחה')

        elif file_type in Music_Types:
            shutil.copyfile(file, self.folder_path + "/" + self.event_name + '/' + 'מוזיקה' + '/' + file_name)
            self.dragndrop_list.addItem(file_name.split(".")[0] + ' - נוסף בהצלחה')

        else:
            shutil.copyfile(file, self.folder_path + "/" + self.event_name + '/' + file_name)
            self.dragndrop_list.addItem(file_name.split(".")[0] + ' - נוסף בהצלחה')

    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ GENERAL FUNCTIONS ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def update_budget(self):
        global current_budget

        record_exist = DataBase.fetch_from_db(self, 'budgets', 'one', self.event_name)
        if record_exist:  # if record not empty
            current_budget.append(self.event_name)

            budget = self.ui.event_budget_text.toPlainText().replace(',', ''). \
                replace(' ', '').replace('-', '').replace('/', '').replace('.', '')

            current_budget.append(budget)
            DataBase.update_db(self, 'budgets')
        else:
            budget = {'event_name': self.event_name,
                      'budget': self.ui.event_budget_text.toPlainText().replace(',', ''). \
                          replace(' ', '').replace('-', '').replace('/', '').replace('.', '')
                      }
            DataBase.insert_into_db(self, 'budgets', budget)

        self.subtitle_budget_status()

    def subtitle_budget_status(self):
        # BUDGET DATA
        budget_info = DataBase.fetch_from_db(self, 'budgets', 'one', self.event_name)
        if budget_info:
            self.budget = budget_info[1]
            self.ui.event_budget_text.setText(make_amount_readable(self.budget))
        else:
            self.budget = 0

        # HEADLINE BUDGET STATUS
        no_status = emoji.emojize('אין סטאטוס כרגע, אפשר לשתות קפה בנחת ' + ':coffee:', use_aliases=True)
        self.ui.budget_status_label.setText(no_status)
        self.ui.budget_status_label.setStyleSheet(u"font: 13pt \"Segoe UI Semibold\";\n"
                                                  "color: rgb(151,186,172);\n"
                                                  "background-color: rgb(255,255,255);\n"
                                                  "border-radius: 15px;")

        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("SELECT * FROM customers WHERE event_name=:event_name AND lead=:lead",
                       {'event_name': self.event_name,
                        'lead': 'לא'
                        })
        self.event_details = cursor.fetchall()

        cursor.execute("SELECT category, price FROM payments WHERE event_name=:event_name",
                       {'event_name': self.event_name
                        })
        payments = cursor.fetchall()

        cursor.close()
        db_connect.close()

        if payments:
            total_payments = 0
            total_gifts = 0
            for payment in payments:
                if payment[1] and payment[0] != 'מתנות':
                    total_payments += float(payment[1])
                if payment[0] == 'מתנות':
                    total_gifts += float(payment[1])
            total_payments -= float(total_gifts)

            if self.budget == '':
                self.budget = 0

            budget_leftover = make_amount_readable(str(int(self.budget) - int(total_payments)))
            over_budget = emoji.emojize('חורג מהתקציב ב' + str(budget_leftover) + ' ש"ח ' + ':astonished:',
                                        use_aliases=True)
            under_budget = emoji.emojize('נשאר עוד ' + str(budget_leftover) + ' ש"ח לבזבז ' + ':relieved:',
                                         use_aliases=True)
            on_budget = emoji.emojize('בול בתקציב ' + ':raised_hands:', use_aliases=True)

            if int(self.budget) - total_payments > 1000:
                self.ui.budget_status_label.setText(under_budget)
                self.ui.budget_status_label.setStyleSheet(u"font: 13pt \"Segoe UI Semibold\";\n"
                                                          "color: rgb(16, 202, 44);\n"
                                                          "background-color: rgb(255,255,255);\n"
                                                          "border-radius: 15px;")
            elif int(self.budget) - total_payments < 0:
                self.ui.budget_status_label.setText(over_budget)
                self.ui.budget_status_label.setStyleSheet(u"font: 13pt \"Segoe UI Semibold\";\n"
                                                          "color: rgb(255, 0, 0);\n"
                                                          "background-color: rgb(255,255,255);\n"
                                                          "border-radius: 15px;")
            else:
                self.ui.budget_status_label.setText(on_budget)
                self.ui.budget_status_label.setStyleSheet(u"font: 13pt \"Segoe UI Semibold\";\n"
                                                          "color: rgb(0, 0, 0);\n"
                                                          "background-color: rgb(255,255,255);\n"
                                                          "border-radius: 15px;")

    def status_color_change(self):
        if self.ui.status_combobox.currentText() == 'בתהליך':
            status_stylesheet = "font: 12pt 'Segoe UI Semilight'; border: 1px solid #090206; " \
                                "background-color: rgb(255,223,128); border-radius: 10px;"
        elif self.ui.status_combobox.currentText() == 'בוצע':
            status_stylesheet = "font: 12pt 'Segoe UI Semilight'; border: 1px solid #090206; " \
                                "background-color: rgb(165,212,106); border-radius: 10px;"
        elif self.ui.status_combobox.currentText() == 'לא רלוונטי':
            status_stylesheet = "font: 12pt 'Segoe UI Semilight'; border: 1px solid #090206; " \
                                "background-color: rgb(227,140,140); border-radius: 10px;"
        else:
            status_stylesheet = "font: 12pt 'Segoe UI Semilight'; border: 1px solid #090206; " \
                                "background-color: rgb(255,255,255); border-radius: 10px;"
        self.ui.status_combobox.setStyleSheet(status_stylesheet)

    def switchto_payments_page(self):
        self.change_btn_color('payments')

        self.ui.payments_btn_line.show()
        self.ui.missions_btn_line.hide()
        self.ui.files_btn_line.hide()
        self.dragndrop_list.hide()

        self.ui.payment_delete_all_warning.hide()
        self.ui.payments_all_select_delete_label.hide()
        self.ui.payments_all_select_delete_no_btn.hide()
        self.ui.payments_all_select_delete_yes_btn.hide()
        self.ui.payment_delete_selection_warning.hide()
        self.ui.payment_single_select_delete_yes_btn.hide()
        self.ui.payment_single_select_delete_no_btn.hide()
        self.ui.payment_single_select_delete_label.hide()

        self.ui.events_stackedWidget.show()
        self.ui.events_stackedWidget.setCurrentWidget(self.ui.payments_page)

    def switchto_missions_page(self):
        self.change_btn_color('missions')

        self.dragndrop_list.hide()

        self.ui.payments_btn_line.hide()
        self.ui.missions_btn_line.show()
        self.ui.files_btn_line.hide()

        self.ui.missions_delete_all_warning.hide()
        self.ui.missions_all_select_delete_label.hide()
        self.ui.missions_all_select_delete_no_btn.hide()
        self.ui.missions_all_select_delete_yes_btn.hide()
        self.ui.mission_delete_selection_warning.hide()
        self.ui.mission_single_select_delete_no_btn.hide()
        self.ui.mission_single_select_delete_yes_btn.hide()
        self.ui.mission_single_select_delete_label.hide()

        self.ui.events_stackedWidget.show()
        self.ui.events_stackedWidget.setCurrentWidget(self.ui.missions_page)

    def switchto_files_page(self):
        self.change_btn_color('files')

        self.ui.payments_btn_line.hide()
        self.ui.missions_btn_line.hide()
        self.ui.files_btn_line.show()
        self.dragndrop_list.show()

        self.ui.events_stackedWidget.show()
        self.ui.events_stackedWidget.setCurrentWidget(self.ui.files_page)

    def show_client_details(self):
        if self.ui.client_details_frame.isHidden():
            # self.ui.client_details_frame.setGeometry(QRect(546, 150, 305, 91))
            if len(self.event_details) == 1:
                event_name_1, event_date_1, contact_name_1, contact_address_1, contact_phone_1, contact_email_1, \
                contract_1, rate_1, location_1, location_address_1, place_1, lead_1 = self.event_details[0]

                self.ui.client_detail_name2.clear()
                self.ui.client_detail_phone2.clear()
                self.ui.client_detail_address2.clear()
                self.ui.client_detail_email2.clear()

            elif len(self.event_details) == 2:
                event_name_1, event_date_1, contact_name_1, contact_address_1, contact_phone_1, contact_email_1, \
                contract_1, rate_1, location_1, location_address_1, place_1, lead_1 = self.event_details[0]
                event_name_2, event_date_2, contact_name_2, contact_address_2, contact_phone_2, contact_email_2, \
                contract_2, rate_2, location_2, location_address_2, place_2, lead_2 = self.event_details[1]
            try:
                self.ui.client_detail_name1.setText(contact_name_1)
                self.ui.client_detail_phone1.setText(contact_phone_1)
                self.ui.client_detail_address1.setText(contact_address_1)
                self.ui.client_detail_email1.setText(contact_email_1)

                self.ui.client_detail_name2.setText(contact_name_2)
                self.ui.client_detail_phone2.setText(contact_phone_2)
                self.ui.client_detail_address2.setText(contact_address_2)
                self.ui.client_detail_email2.setText(contact_email_2)
            except:
                pass

            self.ui.client_details_frame.show()

        else:
            self.ui.client_details_frame.hide()

    def show_event_details(self):
        if self.ui.event_detail_frame.isHidden():
            for data in self.event_details:
                event_name, event_date, contact_name, contact_address, contact_phone, \
                contact_email, contract, rate, location, location_address, place, lead = data

                if event_date == '' or event_date == ' ' or event_date == None:
                    event_date = 'לא נקבע תאריך'
                self.ui.event_detail_name.setText(event_name + ' - ' + event_date)
                if contract == 'קיים':
                    amount = ''
                    if rate:
                        amount = 'ע"ס ' + str(rate) + ' ש"ח'
                    self.ui.event_detail_contract.setText('נחתם חוזה ' + amount)
                else:
                    self.ui.event_detail_contract.setText('עוד לא נחתם חוזה')

                spot = ''
                if location:
                    spot = ', ' + location
                self.ui.event_detail_location.setText(place + spot)
                self.ui.event_detail_location_address.setText(location_address)

            if self.ui.event_budget_text == '':
                self.ui.event_budget_label.hide()
            else:
                self.ui.event_budget_label.show()

            self.ui.event_detail_frame.show()
        else:
            self.ui.event_detail_frame.hide()

    def show_add_buttons(self):
        self.dragndrop_list.hide()

        if self.ui.missions_page_btn.isHidden():
            self.ui.missions_page_btn.show()
            self.ui.missions_page_btn_label.show()
        else:
            self.ui.missions_page_btn.hide()
            self.ui.missions_page_btn_label.hide()
            self.ui.missions_btn_line.hide()

        if self.ui.payments_page_btn.isHidden():
            self.ui.payments_page_btn.show()
            self.ui.payments_page_btn_label.show()
        else:
            self.ui.payments_page_btn.hide()
            self.ui.payments_page_btn_label.hide()
            self.ui.payments_btn_line.hide()

        if self.ui.files_page_btn.isHidden():
            self.ui.files_page_btn.show()
            self.ui.files_page_btn_label.show()
        else:
            self.ui.files_page_btn.hide()
            self.ui.files_page_btn_label.hide()
            self.ui.files_btn_line.hide()

        if self.ui.events_stackedWidget.isHidden():
            pass
        else:
            self.ui.events_stackedWidget.hide()

        self.change_btn_color('default')

    def clear_text(self):
        self.ui.supplier_service_text.clear()
        self.ui.category_combobox.setCurrentIndex(0)
        self.ui.supplier_details_text.clear()
        self.ui.supplier_price_text.clear()
        self.ui.guest_num_input.clear()
        self.ui.supplier_advance_text.clear()
        self.ui.supplier_tip_text.clear()
        self.ui.advance_checkbox.setChecked(0)
        self.ui.suppliers_combobox.setCurrentIndex(0)
        self.ui.payments_excel_warning.hide()

        self.ui.mission_name_text.clear()
        self.ui.mission_due_text.clear()
        self.ui.mission_notes_text.clear()
        self.ui.status_combobox.setCurrentIndex(0)
        self.ui.missions_excel_warning.hide()

        self.ui.service_mendatory_label.hide()
        self.ui.supplier_mendatory_label.hide()
        self.ui.mission_mendatory_label.hide()
        self.ui.mission_name_text.setStyleSheet(u"font: 12pt \"Segoe UI Semilight\";\n"
                                                "border: 3px solid #6248F0;\n"
                                                "background-color: rgb(255,255,255);\n"
                                                "border-radius: 20px;")
        self.ui.supplier_service_text.setStyleSheet(u"font: 12pt \"Segoe UI Semilight\";\n"
                                                    "border: 3px solid #6248F0;\n"
                                                    "background-color: rgb(255,255,255);\n"
                                                    "border-radius: 20px;")
        self.ui.suppliers_combobox.setStyleSheet(u"font: 12pt \"Segoe UI Semilight\";\n"
                                                 "border: 3px solid #6248F0;\n"
                                                 "background-color: rgb(255,255,255);\n"
                                                 "border-radius: 20px;")
        self.ui.client_detail_name1.clear()

    def clear_all(self):
        self.ui.event_folder_label.hide()
        self.ui.event_excel_label.hide()
        self.ui.client_details_frame.hide()
        self.ui.event_detail_frame.hide()

        self.delete_event_check = False
        self.ui.event_delete_label.hide()
        self.ui.delete_warning_label.hide()

        self.ui.payments_page_btn.hide()
        self.ui.payments_page_btn_label.hide()
        self.ui.payments_btn_line.hide()

        self.ui.guest_num_input.hide()

        self.ui.missions_page_btn.hide()
        self.ui.missions_page_btn_label.hide()
        self.ui.missions_btn_line.hide()

        self.ui.files_page_btn.hide()
        self.ui.files_page_btn_label.hide()
        self.ui.files_btn_line.hide()

        self.ui.event_clear_all_btn.hide()
        self.ui.event_collapse_all_btn.show()

        self.ui.missions_excel_warning.hide()
        self.ui.payments_excel_warning.hide()

        self.ui.events_stackedWidget.hide()
        self.dragndrop_list.hide()
        self.clear_text()

    def collapse_all(self):
        if self.ui.client_details_frame.isHidden():
            self.show_client_details()
        if self.ui.event_detail_frame.isHidden():
            self.show_event_details()
        if self.ui.payments_page_btn.isHidden():
            self.show_add_buttons()

        self.ui.event_collapse_all_btn.hide()
        self.ui.event_clear_all_btn.show()

    def pop_window(self, btn_name):
        self.ui.headline_label.setText(btn_name)
        self.event_name = self.ui.headline_label.text()

        db_connect = sqlite3.connect(DB)
        cursor = db_connect.cursor()
        cursor.execute("SELECT * FROM customers WHERE event_name=:event_name AND lead=:lead",
                       {'event_name': self.event_name,
                        'lead': 'לא'
                        })
        self.event_details = cursor.fetchall()

        # HEADLINE DATE LABEL
        new_date = make_date_standard(self.event_details[0][1])

        self.ui.event_date_label.setText(QCoreApplication.translate("EventWindow",
                                                                    u"<html><head/><body><p align=\"center\"><span style=\" font-weight:600;\">\u29bf " + new_date + " \u29bf</span></p></body></html>",
                                                                    None))

        # CATEGORIES DATA
        categories = DataBase.fetch_from_db(self, 'categories')
        for item in categories:
            self.ui.category_combobox.addItem(item[0])

        cursor.close()
        db_connect.close()

        self.subtitle_budget_status()

        self.dragndrop_list.event_name = self.ui.headline_label.text()
        self.dragndrop_list.clear()
        self.dragndrop_list.hide()

        self.ui.missions_page_btn.hide()
        self.ui.missions_page_btn_label.hide()
        self.ui.payments_page_btn.hide()
        self.ui.payments_page_btn_label.hide()
        self.ui.files_page_btn.hide()
        self.ui.files_page_btn_label.hide()

        self.ui.mission_single_select_delete_label.hide()
        self.ui.mission_single_select_delete_no_btn.hide()
        self.ui.mission_single_select_delete_yes_btn.hide()
        self.ui.missions_all_select_delete_label.hide()
        self.ui.missions_all_select_delete_yes_btn.hide()
        self.ui.missions_all_select_delete_no_btn.hide()

        self.ui.edit_mission_btn.hide()
        self.ui.edit_payment_btn.hide()

        self.payments_tree_setup()
        self.missions_tree_setup()
        self.clear_text()

        if not os.path.isdir(self.folder_path + btn_name):
            os.mkdir(self.folder_path + btn_name)
            personal_folders('create', btn_name)

        self.clear_all()
        self.show()

    def pop_alerts(self, btn_name):  # open missions window based on event window
        # startIndex = str(btn_name).find('\"') + 1
        # endIndex = str(btn_name).find('\"', startIndex + 1)
        # btn_name = str(btn_name)[startIndex:endIndex]
        # self.show()
        pass

    def change_btn_color(self, btn):
        if btn == 'payments':
            self.ui.payments_page_btn.setStyleSheet(u"QPushButton {\n"
                                                    "	image: url(:/icons/misc/supplier_payment_icon.png);\n"
                                                    "	background-color: rgb(255,255,255);\n"
                                                    "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                    "	border-radius: 25px;\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:hover {\n"
                                                    "	background-color: rgb(255,255,255);\n"
                                                    "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:pressed {\n"
                                                    "	background-color: rgb(255,255,255);\n"
                                                    "}")
            self.ui.missions_page_btn.setStyleSheet(u"QPushButton {\n"
                                                    "	image: url(:/icons/misc/task-icon.jpg);\n"
                                                    "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                    "	border-radius: 25px;\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:hover {\n"
                                                    "	background-color: rgb(236,211,175);\n"
                                                    "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:pressed {\n"
                                                    "	background-color: rgb(168,230,207);\n"
                                                    "}")
            self.ui.files_page_btn.setStyleSheet(u"QPushButton {\n"
                                                 "	image: url(:/icons/misc/add_files_icon.png);\n"
                                                 "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                 "	border-radius: 25px;\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:hover {\n"
                                                 "	background-color: rgb(236,211,175);\n"
                                                 "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:pressed {\n"
                                                 "	background-color: rgb(168,230,207);\n"
                                                 "}")
        elif btn == 'missions':
            self.ui.payments_page_btn.setStyleSheet(u"QPushButton {\n"
                                                    "	image: url(:/icons/misc/supplier_payment_icon.png);\n"
                                                    "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                    "	border-radius: 25px;\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:hover {\n"
                                                    "	background-color: rgb(236,211,175);\n"
                                                    "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:pressed {\n"
                                                    "	background-color: rgb(168,230,207);\n"
                                                    "}")
            self.ui.missions_page_btn.setStyleSheet(u"QPushButton {\n"
                                                    "	image: url(:/icons/misc/task-icon.jpg);\n"
                                                    "	background-color: rgb(255,255,255);\n"
                                                    "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                    "	border-radius: 25px;\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:hover {\n"
                                                    "	background-color: rgb(255,255,255);\n"
                                                    "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:pressed {\n"
                                                    "	background-color: rgb(255,255,255);\n"
                                                    "}")
            self.ui.files_page_btn.setStyleSheet(u"QPushButton {\n"
                                                 "	image: url(:/icons/misc/add_files_icon.png);\n"
                                                 "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                 "	border-radius: 25px;\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:hover {\n"
                                                 "	background-color: rgb(236,211,175);\n"
                                                 "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:pressed {\n"
                                                 "	background-color: rgb(168,230,207);\n"
                                                 "}")
        elif btn == 'files':
            self.ui.payments_page_btn.setStyleSheet(u"QPushButton {\n"
                                                    "	image: url(:/icons/misc/supplier_payment_icon.png);\n"
                                                    "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                    "	border-radius: 25px;\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:hover {\n"
                                                    "	background-color: rgb(236,211,175);\n"
                                                    "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:pressed {\n"
                                                    "	background-color: rgb(168,230,207);\n"
                                                    "}")
            self.ui.missions_page_btn.setStyleSheet(u"QPushButton {\n"
                                                    "	image: url(:/icons/misc/task-icon.jpg);\n"
                                                    "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                    "	border-radius: 25px;\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:hover {\n"
                                                    "	background-color: rgb(236,211,175);\n"
                                                    "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:pressed {\n"
                                                    "	background-color: rgb(168,230,207);\n"
                                                    "}")
            self.ui.files_page_btn.setStyleSheet(u"QPushButton {\n"
                                                 "	image: url(:/icons/misc/add_files_icon.png);\n"
                                                 "	background-color: rgb(255,255,255);\n"
                                                 "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                 "	border-radius: 25px;\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:hover {\n"
                                                 "	background-color: rgb(255,255,255);\n"
                                                 "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:pressed {\n"
                                                 "	background-color: rgb(255,255,255);\n"
                                                 "}")
        else:
            self.ui.payments_page_btn.setStyleSheet(u"QPushButton {\n"
                                                    "	image: url(:/icons/misc/supplier_payment_icon.png);\n"
                                                    "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                    "	border-radius: 25px;\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:hover {\n"
                                                    "	background-color: rgb(236,211,175);\n"
                                                    "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:pressed {\n"
                                                    "	background-color: rgb(168,230,207);\n"
                                                    "}")
            self.ui.missions_page_btn.setStyleSheet(u"QPushButton {\n"
                                                    "	image: url(:/icons/misc/task-icon.jpg);\n"
                                                    "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                    "	border-radius: 25px;\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:hover {\n"
                                                    "	background-color: rgb(236,211,175);\n"
                                                    "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                    "}\n"
                                                    "\n"
                                                    "QPushButton:pressed {\n"
                                                    "	background-color: rgb(168,230,207);\n"
                                                    "}")
            self.ui.files_page_btn.setStyleSheet(u"QPushButton {\n"
                                                 "	image: url(:/icons/misc/add_files_icon.png);\n"
                                                 "	font: 63 20pt \"Segoe UI Semilight\";\n"
                                                 "	border-radius: 25px;\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:hover {\n"
                                                 "	background-color: rgb(236,211,175);\n"
                                                 "	font: 63 20pt \"Segoe UI Semibold\";\n"
                                                 "}\n"
                                                 "\n"
                                                 "QPushButton:pressed {\n"
                                                 "	background-color: rgb(168,230,207);\n"
                                                 "}")

    def export_to_excel(self):
        # try:
        self.ui.missions_excel_warning.hide()
        self.ui.payments_excel_warning.hide()
        btn_clicked = self.sender()

        create_excel(btn_clicked, self.event_name)

        # except Exception as e:

        # self.ui.missions_excel_warning.show()
        # self.ui.payments_excel_warning.show()

    def delete_event(self):
        self.ui.delete_warning_label.show()
        if self.delete_event_check:
            self.ui.delete_warning_label.hide()
            for data in self.event_details:
                event_name, event_date, contact_name, contact_address, contact_phone, \
                contact_email, contract, rate, location, location_address, place, lead = data
                DataBase.delete_from_db(self, 'customers', 'one', self.event_name, contact_name)
            personal_folders('delete', event_name)
            self.close()

        else:
            self.delete_event_check = False

        self.delete_event_check = True


class WarningWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_WarningWindow()
        self.ui.setupUi(self)

        self.hide()

    def pop_window(self, *args):
        self.show()
        self.raise_()


# SPLASH SCREEN
class SplashScreen(QMainWindow):
    singleton: 'SplashScreen' = None

    def __init__(self):
        super().__init__()
        self.ui = Ui_SplashScreen()
        self.ui.setupUi(self)
        self.setWindowIcon(QtGui.QIcon('Pheonix.ico'))

        LoadDB = DataBase()

        # CREATE DB TABLES
        LoadDB.create_tables()
        # ACTIVATE DB BACKUP
        LoadDB.run_DB_backup()
        # FETCH EVENTS AND LEADS
        LoadDB.fetch_events_and_leads()

        ## REMOVE TITLE BAR
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        ## DROP SHADOW EFFECT
        self.shadow = QGraphicsDropShadowEffect(self)
        self.shadow.setBlurRadius(20)
        self.shadow.setXOffset(0)
        self.shadow.setYOffset(0)
        self.shadow.setColor(QColor(0, 0, 0, 60))
        self.ui.main_frame.setGraphicsEffect(self.shadow)

        ## QTIMER ==> START
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.progress)
        # TIMER IN MILLISECONDS
        self.timer.start(15)

        ## SHOW ==> MAIN WINDOW
        self.show()

    ## ==> APP FUNCTIONS
    def progress(self):
        global counter

        # SET VALUE TO PROGRESS BAR
        self.ui.splash_progress.setValue(counter)

        # CLOSE SPLASH SCREE AND OPEN APP
        if counter > 100:
            # STOP TIMER
            self.timer.stop()

            # SHOW MAIN WINDOW
            self.main = MainWindow()
            self.main.show()

            # CLOSE SPLASH SCREEN
            self.close()

        # INCREASE COUNTER
        counter += 1


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SplashScreen()
    sys.exit(app.exec_())
